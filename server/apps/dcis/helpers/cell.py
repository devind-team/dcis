"""Вспомогательный модуль для расчета формул."""
from collections import defaultdict
from itertools import groupby
from typing import Iterable, TypedDict

from django.db.models import Q, QuerySet
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string, get_column_letter
from xlsx_evaluate import Evaluator, Model, ModelCompiler
from xlsx_evaluate.functions.xlerrors import (
    DivZeroExcelError, NaExcelError, NameExcelError,
    NullExcelError, NumExcelError, RefExcelError,
    ValueExcelError,
)

from apps.dcis.models import Cell, Document, Sheet, Value
from .sheet_formula_cache import SheetFormulaContainerCache
from ..models.sheet import KindCell


def get_dependency_cells(
    sheet_containers: list[SheetFormulaContainerCache],
    values: list[Value]
) -> tuple[list[str], list[str], list[str]]:
    """Получаем связанные ячейки.

    Возвращается кортеж:
        - dependency - список зависимых ячеек, даже если там формула, она забирается как значение;
        - inversion - список связных ячеек, которые нужно пересчитать;
        - sheet_containers - последовательность расчета листов.
    """
    dependency_cells: list[str] = []
    inversion_cells: list[str] = []
    sequence_evaluate: list[str] = []
    cells = [f'{value.sheet.name}!{get_column_letter(value.column.index)}{value.row.index}' for value in values]

    while cells:
        cell = cells.pop()
        sheet_name, column_letter, row_index = parse_coordinate(cell)
        sequence_evaluate.append(sheet_name)
        sheet_container: SheetFormulaContainerCache
        for sheet_container in sheet_containers:
            coordinate: str = f'{column_letter}{row_index}' if sheet_name == sheet_container.sheet_name else cell
            inversions: list[str] = [
                normalize_coordinate(coord, sheet_container.sheet_name)
                for coord in sheet_container.inversion.get(coordinate, [])
            ]
            dependency: list[str] = [
                normalize_coordinate(coord, sheet_container.sheet_name)
                for coord in sheet_container.dependency.get(coordinate, {}).keys()
            ]
            dependency_cells.extend(dependency)
            inversion_cells.extend(inversions)
            cells.extend(inversions)
    return dependency_cells, inversion_cells, [el for el, _ in groupby(sequence_evaluate)]


def resolve_cells(
    sheets: Iterable[Sheet],
    document: Document,
    cells: set[str]
) -> tuple[QuerySet[Cell], QuerySet[Value]]:
    """Получаем строки в зависимости от координат."""
    sheet_mapping: dict[str, int] = {sheet.name: sheet.pk for sheet in sheets}
    sheet_cells: defaultdict[int, list[tuple[int, int]]] = defaultdict(list)
    for cell in cells:
        sheet_name, column_letter, row = parse_coordinate(cell)
        sheet_cells[sheet_mapping[sheet_name]].append((column_index_from_string(column_letter), row,))

    cells_query_filter = Q()
    values_query_filter = Q()
    for sheet_id, coordinates in sheet_cells.items():
        for column_index, row_index in coordinates:
            cells_query_filter |= Q(
                column__sheet_id=sheet_id,
                column__index=column_index,
                row__index=row_index,
                row__parent__isnull=True
            )
            values_query_filter |= Q(
                column__sheet_id=sheet_id,
                document=document,
                column__index=column_index,
                row__index=row_index
            )
    related: list[str] = ['row', 'column', 'column__sheet']
    cells: QuerySet[Cell] = Cell.objects.filter(cells_query_filter) \
        .select_related(*related).all()
    values: QuerySet[Value] = Value.objects.filter(values_query_filter)\
        .select_related(*related).all()
    return cells, values


class ValueState(TypedDict):
    """Результативное состояние для расчета новых значений по формулам.

        - value - значение;
        - error - ошибка вычисления формулы;
        - formula - формула;
        - cell - ячейка
    """
    value: str | float | None
    error: str | None
    formula: str | None
    cell: Cell


def resolve_evaluate_state(
    cells: Iterable[Cell],
    values: Iterable[Value],
    inversion_cells: list[str]
) -> dict[str, ValueState]:
    """Строим массив для расчета.

        - cells - массив ячеек для расчета;
        - values - массив уже существующих значений, для расчета новых значений ячейки;
        - inversion_cells - ячейки от которых зависит расчет.
    """
    state: dict[str, ValueState] = {}
    values_state: dict[str, str] = {get_coordinate(v.column.sheet, v): v.value.strip() for v in values}
    cell: Cell
    for cell in cells:
        coord = get_coordinate(cell.column.sheet, cell)
        value = values_state.get(coord, cell.default)
        if (cell.kind == KindCell.NUMERIC or cell.kind == KindCell.FORMULA) and value is not None:
            try:
                value = float(value)
            except ValueError:
                pass
        state[coord]: ValueState = {
            'value': value,
            'error': None,
            'formula': cell.formula if cell.formula and coord in inversion_cells else None,
            'cell': cell
        }
    return state


def evaluate_state(state: dict[str, ValueState], sequence_evaluate: list[str]):
    """Рассчитываем новые значения определенным образом.

    Для текущего листа, например, "Лист1"
      - все значения этого листа приводим к относительному виду
        - Лист1!B2 -> B2
        - Лист2!С4 -> Лист2!С4
      - все формулы не из текущего листа заменяем значениями
      - все формулы текущего листа рассчитываем и обновляем результаты
    """
    for sheet_name in sequence_evaluate:
        input_state: dict[str, str | int | float] = {}
        cell_name: str
        cell_state: ValueState
        for cell_name, cell_state in state.items():
            sn, column, row = parse_coordinate(cell_name)
            if sn == sheet_name:
                if cell_state['formula']:
                    input_state[f'{column}{row}'] = cell_state['formula']
                elif cell_state['value'] is not None:
                    input_state[f'{column}{row}'] = cell_state['value']
            elif cell_state['value'] is not None:
                input_state[cell_name] = cell_state['value']
        compiler: ModelCompiler = ModelCompiler()
        model: Model = compiler.read_and_parse_dict(input_dict=input_state, default_sheet=sheet_name)
        evaluator = Evaluator(model)
        for coordinate in model.formulae:
            success, value = evaluate_formula(evaluator, coordinate)
            state[coordinate]['value'] = value if success else ''
            state[coordinate]['error'] = value if not success else None
    return state


def get_coordinate(sheet: Sheet, vc: Value | Cell) -> str:
    return f'{sheet.name}!{get_column_letter(vc.column.index)}{vc.row.index}'


def parse_coordinate(coord: str, default_sheet: str | None = None) -> tuple[str | None, str, int]:
    """Разбираем координаты."""
    coordinate = coord.split('!')
    if len(coordinate) == 1:
        column, row = coordinate_from_string(coordinate[0])
        return default_sheet, column, row
    column, row = coordinate_from_string(coordinate[1])
    return coordinate[0], column, row


def normalize_coordinate(coord: str, sheet_name: str | None = None) -> str:
    """Нормализация координаты, приведение к виду
        - A1 -> ИмяПоУмолчанию!A1
        - Лист1!B2 -> Лист1!B2
    """
    sheet_name, column, row = parse_coordinate(coord, sheet_name)
    return f'{sheet_name}!{column}{row}'


def evaluate_formula(evaluator: Evaluator, coordinate: str) -> tuple[bool, str]:
    """Вычисление формулы с возвратом возможной ошибки.

    Возвращает (успех, рассчитанное значение или сообщение ошибки).
    """
    try:
        value = evaluator.evaluate(coordinate)
        value_type = type(value)
        if value_type in EXCEL_ERRORS_MAP:
            return False, EXCEL_ERRORS_MAP[value_type]
        return True, str(value)
    except RuntimeError as e:
        if 'Cycle detected' in str(e):
            return False, 'Циклическая ссылка'
        raise e


EXCEL_ERRORS_MAP = {
    NullExcelError: 'Неверный диапазон или диапазоны',
    DivZeroExcelError: 'Деление на 0',
    ValueExcelError: 'Неверный тип операнда или аргумента функции',
    RefExcelError: 'Недействительная ссылка',
    NameExcelError: 'Неверное имя',
    NumExcelError: 'Недопустимые числовые значения',
    NaExcelError: 'Не удалось найти данные',
}
