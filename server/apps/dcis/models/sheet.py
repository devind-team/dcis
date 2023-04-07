from django.db import models
from model_clone import CloneMixin
from openpyxl.utils.cell import get_column_letter

from apps.core.models import User
from .document import Document, Sheet


def get_default_border():
    """Стиль и цвет границ по умолчанию."""
    return {
        'top': None,
        'bottom': None,
        'left': None,
        'right': None,
        'diagonal': None
    }


class Style(models.Model):
    """Абстрактная модель применяемых стилей для страниц сборов."""

    LEFT = 'left'
    CENTER = 'center'
    RIGHT = 'right'

    KIND_HORIZONTAL_ALIGN = (
        (LEFT, 'left'),
        (CENTER, 'center'),
        (RIGHT, 'right')
    )

    TOP = 0
    BOTTOM = 1
    MIDDLE = 2

    KIND_VERTICAL_ALIGN = (
        (TOP, 'top'),
        (MIDDLE, 'middle'),
        (BOTTOM, 'bottom'),
    )

    SINGLE = 'single'
    DOUBLE = 'double'
    SINGLE_ACCOUNTING = 'single_accounting'
    DOUBLE_ACCOUNTING = 'double_accounting'

    KIND_UNDERLINE = (
        (SINGLE, 'single'),
        (DOUBLE, 'double'),
        (SINGLE_ACCOUNTING, 'single_accounting'),
        (DOUBLE_ACCOUNTING, 'double_accounting')
    )

    horizontal_align = models.CharField(
        max_length=10,
        default=None,
        null=True,
        choices=KIND_HORIZONTAL_ALIGN,
        help_text='Горизонтальное выравнивание'
    )
    vertical_align = models.CharField(
        max_length=10,
        default=None,
        null=True,
        choices=KIND_VERTICAL_ALIGN,
        help_text='Вертикальное выравнивание'
    )
    size = models.PositiveIntegerField(default=12, help_text='Размер шрифта')
    strong = models.BooleanField(default=False, help_text='Жирный шрифт')
    italic = models.BooleanField(default=False, help_text='Курсив')
    strike = models.BooleanField(default=False, help_text='Зачеркнутый')
    underline = models.CharField(
        max_length=20,
        default=None,
        null=True,
        choices=KIND_UNDERLINE,
        help_text='Тип подчеркивания'
    )
    color = models.CharField(max_length=16, default='#000000', help_text='Цвет текста')
    background = models.CharField(max_length=16, default='#FFFFFF', help_text='Цвет фона')
    border_style = models.JSONField(default=get_default_border, help_text='Стили границ')
    border_color = models.JSONField(default=get_default_border, help_text='Цвет границ')

    class Meta:
        abstract = True


class KindCell(models.Model):
    """Классификация ячеек."""

    # Формат из openpyxl
    NUMERIC = 'n'
    STRING = 's'
    FORMULA = 'f'
    BOOL = 'b'
    INLINE = 'inlineStr'
    ERROR = 'e'
    FORMULA_CACHE_STRING = 'str'
    DATE = 'd'

    # Дополнительный набор
    TIME = 'time'
    TEXT = 'text'
    MONEY = 'money'
    BIG_MONEY = 'bigMoney'
    FILES = 'fl'

    # Поля из базы данных
    USER = 'user'
    DEPARTMENT = 'department'
    ORGANIZATION = 'organization'
    CLASSIFICATION = 'classification'

    KIND_VALUE = (
        (NUMERIC, 'n'),
        (STRING, 's'),
        (FORMULA, 'f'),
        (BOOL, 'b'),
        (INLINE, 'inlineStr'),
        (ERROR, 'e'),
        (FORMULA_CACHE_STRING, 'str'),
        (DATE, 'd'),
        (TIME, 'time'),
        (TEXT, 'text'),
        (MONEY, 'money'),
        (BIG_MONEY, 'bigMoney'),
        (FILES, 'fl'),
        (USER, 'user'),
        (DEPARTMENT, 'department'),
        (ORGANIZATION, 'organization'),
        (CLASSIFICATION, 'classification'),
    )

    kind = models.CharField(
        max_length=30,
        default=STRING,
        choices=KIND_VALUE,
        help_text='Тип значения'
    )

    class Meta:
        abstract = True


class ColumnDimension(KindCell, models.Model, CloneMixin):
    """Модель стилей для колонки таблицы.

    Ссылка на оригинальный класс из openpyxl:
    https://foss.heptapod.net/openpyxl/openpyxl/-/blob/branch/3.0/openpyxl/worksheet/dimensions.py
    """

    index = models.PositiveIntegerField(help_text='Индекс колонки')
    width = models.PositiveIntegerField(null=True, help_text='Ширина колонки')
    fixed = models.BooleanField(default=False, help_text='Фиксация колонки')
    hidden = models.BooleanField(default=False, help_text='Скрытие колонки')

    created_at = models.DateTimeField(auto_now_add=True, help_text='Дата добавления')
    updated_at = models.DateTimeField(auto_now=True, help_text='Дата обновления')

    sheet = models.ForeignKey(Sheet, on_delete=models.CASCADE, help_text='Лист')
    user = models.ForeignKey(User, null=True, on_delete=models.SET_NULL, help_text='Пользователь')

    class Meta:
        ordering = ('index', 'id',)
        unique_together = [['index', 'sheet']]


class RowDimension(models.Model, CloneMixin):
    """Модель стилей для строки таблицы.

    Кроме того, что таблица плоская, могут быть еще промежуточные агрегаций,
    они реализуются с помощью дополнительных полей.

    - dynamic - описывает динамическую строку
    - aggregation - способ агрегации дочерних строк (SUM, MIN, MAX, AVG)
    - parent - ссылка на родительскую строку
    - document - ссылка на документ.
        Это поле необходимо для того, чтобы к обычным строкам плоской таблицы
        добавлять динамические строки, которые не относятся к определенному документу,
        однако, должны участвовать при заполнении финальной версии документа.
    """
    SUM = 'SUM'
    MIN = 'MIN'
    MAX = 'MAX'
    AVG = 'AVG'

    KIND_AGGREGATION = (
        (SUM, 'sum'),
        (MIN, 'min'),
        (MAX, 'max'),
        (AVG, 'avg'),
    )

    index = models.PositiveIntegerField(help_text='Индекс строки')
    height = models.PositiveIntegerField(null=True, help_text='Высота строки')
    fixed = models.BooleanField(default=False, help_text='Фиксация строки')
    hidden = models.BooleanField(default=False, help_text='Скрытие строки')

    dynamic = models.BooleanField(default=False, help_text='Динамическая ли строка')
    object_id = models.PositiveIntegerField(null=True)
    aggregation = models.CharField(
        max_length=3,
        null=True,
        default=None,
        choices=KIND_AGGREGATION,
        help_text='Агрегирование перечисление (мин, макс) для динамических строк'
    )

    created_at = models.DateTimeField(auto_now_add=True, help_text='Дата добавления')
    updated_at = models.DateTimeField(auto_now=True, help_text='Дата обновления')

    sheet = models.ForeignKey(Sheet, on_delete=models.CASCADE, help_text='Лист')

    parent = models.ForeignKey(
        'self',
        default=None,
        null=True,
        on_delete=models.CASCADE,
        help_text='Родительское правило'
    )
    document = models.ForeignKey(
        Document,
        default=None,
        null=True,
        on_delete=models.CASCADE,
        help_text='Документ, для динамических строк'
    )
    user = models.ForeignKey(User, null=True, on_delete=models.SET_NULL, help_text='Пользователь')

    class Meta:
        ordering = ('index', 'id',)


class Cell(Style, KindCell, models.Model, CloneMixin):
    """Модель ячейки."""
    AGGREGATION_SUM = 'sum'
    AGGREGATION_AVG = 'avg'
    AGGREGATION_MIN = 'min'
    AGGREGATION_MAX = 'max'

    KIND_AGGREGATION = (
        (AGGREGATION_SUM, 'sum'),
        (AGGREGATION_AVG, 'avg'),
        (AGGREGATION_MIN, 'min'),
        (AGGREGATION_MAX, 'max'),
    )

    TEMPLATE_FIELD = [KindCell.STRING, KindCell.TEXT]

    editable = models.BooleanField(default=True, help_text='Редактируемая ячейка')
    formula = models.TextField(null=True, help_text='Формула')
    number_format = models.TextField(null=True, help_text='Форматирование чисел')
    comment = models.TextField(null=True, help_text='Комментарий')
    default = models.TextField(null=True, help_text='Значение по умолчанию')
    default_error = models.TextField(null=True, help_text='Значение ошибки по умолчанию')
    mask = models.TextField(null=True, help_text='Маска для ввода значений')
    tooltip = models.TextField(null=True, help_text='Подсказка')
    is_template = models.BooleanField(default=False, help_text='Является ли поле шаблоном')
    aggregation = models.CharField(
        max_length=5,
        choices=KIND_AGGREGATION,
        null=True,
        default=None,
        help_text='Механизм агрегации'
    )

    column = models.ForeignKey(ColumnDimension, on_delete=models.CASCADE, help_text='Колонка')
    row = models.ForeignKey(RowDimension, on_delete=models.CASCADE, help_text='Строка')

    cells = models.ManyToManyField(
        'self',
        through='RelationshipCells',
        symmetrical=False,
        related_name='related_cells'
    )

    class Meta:
        unique_together = (('column', 'row'),)
        indexes = [
            models.Index(fields=['column'])
        ]

    @property
    def is_aggregation(self) -> bool:
        """Является ли ячейка агрегационной."""
        return self.aggregation is not None


class RelationshipCells(models.Model):
    """Модель много ко многим для ячеек."""

    from_cell = models.ForeignKey(Cell, related_name='from_cells', on_delete=models.CASCADE)
    to_cell = models.ForeignKey(Cell, related_name='to_cells', on_delete=models.CASCADE)


class MergedCell(models.Model, CloneMixin):
    """Модель объединенной ячейки."""

    min_col = models.IntegerField(help_text='Начальная позиция в колонке')
    min_row = models.IntegerField(help_text='Начальная позиция в строке')
    max_col = models.IntegerField(help_text='Конечная позиция в колонке')
    max_row = models.IntegerField(help_text='Конечная позиция в строке')

    sheet = models.ForeignKey(Sheet, on_delete=models.CASCADE, help_text='Лист')

    class Meta:
        ordering = ('min_col', 'min_row',)

    def __str__(self) -> str:
        return f'{get_column_letter(self.min_col)}{self.min_row}:' \
               f'{get_column_letter(self.max_col)}{self.max_row}'

    @property
    def colspan(self) -> int:
        return self.max_col - self.min_col + 1

    @property
    def rowspan(self) -> int:
        return self.max_row - self.min_row + 1

    @property
    def target(self) -> str:
        return f'{get_column_letter(self.min_col)}{self.min_row}'

    @property
    def cells(self) -> list[str]:
        not_cell: list[str] = []
        for row in range(self.min_row, self.max_row + 1):
            for col in range(self.min_col, self.max_col + 1):
                if col == 1 and row == 1:
                    continue
                not_cell.append(f'{get_column_letter(col)}{row}')
        return not_cell


class Value(models.Model, CloneMixin):
    """Модель значения поля таблицы.

    Привязка значения осуществляется к дивизиону через документ,
    но в документе может быть несколько листов.
    В связи с этим добавлено поле sheet, которое точно определяет к какому листу имеет отношение таблицы.
    """

    value = models.TextField(help_text='Значение')
    payload = models.JSONField(null=True, help_text='Дополнительные данные')
    error = models.CharField(max_length=255, null=True, help_text='Текст ошибки')

    document = models.ForeignKey(Document, on_delete=models.CASCADE, help_text='Документ')
    sheet = models.ForeignKey(Sheet, on_delete=models.CASCADE, help_text='Лист')

    column = models.ForeignKey(ColumnDimension, on_delete=models.CASCADE, help_text='Колонка')
    row = models.ForeignKey(RowDimension, on_delete=models.CASCADE, help_text='Строка')

    class Meta:
        indexes = [
            models.Index(fields=['document', 'sheet']),
            models.Index(fields=['document', 'column'])
        ]
