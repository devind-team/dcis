"""Модуль, отвечающий за обновление с документов."""

import posixpath
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from typing import Iterable

from django.conf import settings
from django.db.models import Q, QuerySet
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from apps.core.models import User
from apps.dcis.models import Cell, ColumnDimension, Document, MergedCell, Period, RowDimension, Sheet, Value
from apps.dcis.permissions import can_view_document
from apps.dcis.services.document_services import get_document_sheets


@dataclass
class BuildRow:
    """Дата класс содержащий строку и основную информацию о строке."""
    row: RowDimension
    row_add_date: str
    row_update_date: str
    division_name: str
    division_head: str
    user: str


@dataclass
class BuildCell:
    """Дата класс содержащий собираемую информацию о ячейки."""
    cell: Cell
    value: str
    alignment: Alignment
    font: Font
    border: Border
    pattern_fill: PatternFill


class DocumentUnload:
    """Выгрузка документа в формате Excel."""

    ALLOW_ADDITIONAL: list[str] = ['row_add_date', 'row_update_date', 'division_name', 'division_head', 'user']
    DIVISION_INFO_CACHE: dict[int, tuple[str, str]] = {}

    def __init__(self, document: Document, additional: list[str]) -> None:
        """Инициализация.
        - document - выгружаемый документ
        - host - текущий хост
        - additional - дополнительные параметры
        """
        self.document = document
        self.period = Period.objects.select_related('project').get(pk=self.document.period_id)
        self.project = self.period.project
        self.sheets: QuerySet[Sheet] = get_document_sheets(document).prefetch_related(
            'columndimension_set', 'mergedcell_set'
        ).all()
        self.additional = [field for field in additional if field in self.ALLOW_ADDITIONAL]
        self.path = settings.DOCUMENTS_DIR / f'document_{datetime.now().strftime("%d-%m-%Y_%H-%M-%S")}.xlsx'

    def unload(self) -> str:
        """Выгрузка."""
        wb = Workbook()
        wb.remove(wb.active)
        for sheet in self.sheets:
            ws: Worksheet = wb.create_sheet(sheet.name)
            columns = sheet.columndimension_set.all()
            rows = sheet.rowdimension_set.filter(
                Q(parent__isnull=True) | Q(document=self.document, parent_id__isnull=False)
            )
            rows_id = [row.id for row in rows]
            cells = Cell.objects.filter(row_id__in=rows_id)
            values = Value.objects.filter(document=self.document, row_id__in=rows_id)
            build_rows = self._build_rows(rows)
            build_cells = self._build_cells(cells, values)
            build_merged_cells = self._build_merged_cells(sheet.mergedcell_set.all())
            self._save_rows(ws, columns, build_rows, build_cells, build_merged_cells)
            for column in columns:
                if column.width:
                    ws.column_dimensions[get_column_letter(column.index)].width = column.width // 7
            for ci in range(len(columns) + 1, len(columns) + len(self.additional) + 1):
                ws.column_dimensions[get_column_letter(ci)].width = 20

        wb.save(self.path)
        return posixpath.relpath(self.path, settings.BASE_DIR)

    def _save_rows(
        self,
        ws: Worksheet,
        columns: Iterable[ColumnDimension],
        build_rows: list[BuildRow],
        build_cells: dict[str, BuildCell],
        build_merged_cells: dict[int, list[MergedCell]],
    ):
        """Сохранение строк в xlsx файл."""
        offset_row: int = 0
        for row_index, build_row in enumerate(build_rows, 1):
            self._save_cells(ws, columns, build_cells, build_row, row_index)

            # Пропускаем объединение, если строка дочерняя
            if build_row.row.parent_id is not None:
                offset_row += 1

            # Высота строки, если она задана
            if build_row.row.height:
                ws.row_dimensions[row_index].height = build_row.row.height

            # Объединяем ячейки
            if row_index - offset_row in build_merged_cells:
                for merge_cell in build_merged_cells[row_index - offset_row]:
                    ws.merge_cells(
                        start_column=merge_cell.min_col,
                        start_row=merge_cell.min_row + offset_row,
                        end_column=merge_cell.max_col,
                        end_row=merge_cell.max_row + offset_row,
                    )

    def _save_cells(
        self,
        ws: Worksheet,
        columns: Iterable[ColumnDimension],
        build_cells: dict[str, BuildCell],
        build_row: BuildRow,
        row_index: int,
    ) -> None:
        """Сохранение ячеек в xlsx файл."""
        # Основные колонки
        column_index = 1
        for column in columns:
            cell: BuildCell | None = build_cells.get(f'{column.pk}:{build_row.row.pk}')
            if cell:
                ws.cell(row_index, column_index, cell.value)
                ws.cell(row_index, column_index).alignment = cell.alignment
                ws.cell(row_index, column_index).font = cell.font
                ws.cell(row_index, column_index).border = cell.border
                ws.cell(row_index, column_index).fill = cell.pattern_fill
            column_index += 1

        # Дополнительные колонки
        if self.project.division_name == 'organization':
            if row_index == 1:
                ws.cell(row_index, column_index, self.document.updated_at.strftime('%d.%m.%Y-%H:%M:%S'))
        else:
            for ac in self.additional:
                ws.cell(row_index, column_index, getattr(build_row, ac))
                column_index += 1

    def _build_rows(self, rows: Iterable[RowDimension], parent_id: int | None = None) -> list[BuildRow]:
        """Функция собирает все строки, включая дочерние в плоский массив."""
        date_format: str = '%H:%M %d.%m.%Y'
        build_rows: list[BuildRow] = []
        current_rows = [row for row in rows if row.parent_id == parent_id]
        for current_row in current_rows:
            division_name, division_head = self._division_info(current_row.user)
            build_row = BuildRow(
                current_row,
                current_row.created_at.strftime(date_format),
                current_row.updated_at.strftime(date_format),
                division_name,
                division_head,
                current_row.user.get_full_name if current_row.user is not None else '',
            )
            build_rows = [*build_rows, build_row, *self._build_rows(rows, current_row.pk)]
        return build_rows

    @staticmethod
    def _build_merged_cells(merged_cells: Iterable[MergedCell]) -> dict[int, list[MergedCell]]:
        """Сбор объединенных ячеек в словарь по последней строке"""
        build_mc: dict[int, list[MergedCell]] = defaultdict(list)
        for merge_cell in merged_cells:
            build_mc[merge_cell.max_row].append(merge_cell)
        return build_mc

    def _division_info(self, user: User | None) -> tuple[str, str]:
        """Функция возвращает название дивизиона и начальника этого дивизиона.

        Возвращать информацию можно только из моделей для которых указаны поля:
            - user - начальник дивизиона
            - users - список сотрудников дивизиона
        Модель должна содержать:
            - name - название дивизиона
            - code - код дивизиона
         related_name для пользователя должно подчиняться протоколу
            getattr(user, project.content_type.model) -> получаем дивизион, где пользователь начальник
            getattr(user, f'{project.content_type.model}s') -> получаем мой список дивизионов
        """
        if user is None or self.project.content_type.model not in ['department']:
            return '', ''
        if user.id in self.DIVISION_INFO_CACHE:
            return self.DIVISION_INFO_CACHE[user.id]
        divisions = getattr(user, f'{self.project.content_type.model}s').all()
        division_name = ', '.join([f'{division.name} ({division.code})' for division in divisions])
        division_head = ', '.join([division.user.get_full_name for division in divisions if division.user is not None])
        self.DIVISION_INFO_CACHE[user.id] = division_name, division_head,
        return division_name, division_head

    def _build_cells(self, cells: Iterable[Cell], values: Iterable[Value]) -> dict[str, BuildCell]:
        """Собираем ячейки в словарь для индексации."""
        build_values: dict[str, Value] = {f'{value.column_id}:{value.row_id}': value.value for value in values}
        return {
            f'{cell.column_id}:{cell.row_id}': BuildCell(
                cell,
                build_values.get(f'{cell.column_id}:{cell.row_id}', cell.default),
                self._cell_alignment(cell),
                self._cell_font(cell),
                self._cell_border(cell),
                self._cell_pattern_fill(cell),
            )
            for cell in cells
        }

    @staticmethod
    def _cell_alignment(cell: Cell) -> Alignment:
        """Получение выравнивания ячейки."""
        return Alignment(
            vertical=cell.vertical_align if cell.vertical_align != 'middle' else 'center',
            horizontal=cell.horizontal_align,
            wrap_text=True,
        )

    @staticmethod
    def _cell_font(cell: Cell) -> Font:
        """Получение шрифта ячейки."""
        return Font(
            size=cell.size,
            bold=cell.strong,
            italic=cell.italic,
            strike=cell.strike,
            underline=cell.underline,
            color=f'{cell.color[1:]}',
        )

    def _cell_border(self, cell: Cell) -> Border:
        """Получение границы ячейки."""
        border_styles = {
            position: self._cell_border_side(cell, position)
            for position in ['top', 'bottom', 'left', 'right', 'diagonal']
        }
        return Border(
            diagonalDown=cell.border_style.get('diagonalDown'),
            diagonalUp=cell.border_style.get('diagonalUp'),
            **border_styles
        )

    @staticmethod
    def _cell_pattern_fill(cell: Cell) -> PatternFill:
        """Получение паттерна заливки ячейки."""
        return PatternFill(
            fill_type='solid',
            start_color=f'{cell.background[1:]}',
            end_color=f'{cell.background[1:]}'
        )

    @staticmethod
    def _cell_border_side(cell: Cell, position: str) -> Side:
        """Получение настроек границы ячейки."""
        return Side(
            border_style=cell.border_style.get(position),
            color=f'{cell.border_color[position][1:]}' if cell.border_color[position] else None
        )


def document_upload(user: User, document_id: int, additional: list[str] | None = None) -> str:
    """Функция выгрузки документа."""
    if not additional:
        additional = []
    document = Document.objects.get(pk=document_id)
    can_view_document(user, document)
    document_unload = DocumentUnload(document, additional)
    return document_unload.unload()
