from dataclasses import asdict, dataclass, field
from pathlib import PosixPath
from typing import Iterator

from django.core.files.base import File
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell as OpenpyxlCell, TYPE_FORMULA
from openpyxl.styles.colors import COLOR_INDEX, Color, WHITE
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string, get_column_letter
from openpyxl.worksheet.dimensions import (
    ColumnDimension as OpenpyxlColumnDimension,
    DimensionHolder,
    RowDimension as OpenpyxlRowDimension,
)
from openpyxl.worksheet.merge import MergeCell as OpenpyxlMergedCell
from xlsx_evaluate import Evaluator, ModelCompiler

from apps.dcis.helpers.cell import evaluate_formula
from apps.dcis.helpers.sheet_formula_cache import SheetFormulaContainerCache
from apps.dcis.helpers.theme_to_rgb import theme_and_tint_to_rgb
from apps.dcis.models import Cell, ColumnDimension, MergedCell, Period, RowDimension, Sheet
from apps.dcis.models.sheet import KindCell


@dataclass
class BuildStyle:
    """Описание стилей"""
    horizontal_align: str
    vertical_align: str
    size: float
    strong: bool
    italic: bool
    strike: bool
    underline: bool
    color: str
    background: str
    border_style: dict[str, str]


@dataclass
class BuildColumnDimension:
    """Построение колонки."""
    width: int
    hidden: bool
    style: BuildStyle


@dataclass
class BuildRowDimension:
    """Построение строки."""
    height: int | None
    style: BuildStyle


@dataclass
class BuildCell(BuildStyle):
    """Построение ячеек."""
    column_id: int
    row_id: int
    kind: str
    editable: bool
    coordinate: str | None = None
    formula: str | None = None
    number_format: str | None = None
    comment: str | None = None
    default: str | None = None
    default_error: str | None = None
    border_color: dict[str, str] = None


@dataclass
class BuildMergedCell:
    """Построение объединенных ячеек."""
    min_col: int
    min_row: int
    max_col: int
    max_row: int


@dataclass
class BuildSheet:
    name: str
    columns_dimension: dict[int, BuildColumnDimension]
    fixed_column: int | None
    rows_dimension: dict[int, BuildRowDimension]
    fixed_row: int | None
    cells: list[BuildCell]
    merged_cells: list[BuildMergedCell]

    cache_container: SheetFormulaContainerCache = field(init=False)

    def __post_init__(self):
        self.cache_container = SheetFormulaContainerCache(self.name)


class ExcelExtractor:
    """Парсинг xlsx файла в структуру данных для последовательной загрузки в базу данных."""

    def __init__(self, filename: PosixPath | File, readonly_fill_color: bool) -> None:
        """Инициализация.

        :param filename - путь к файлу Excel или файл Excel
        :param readonly_fill_color - запретить редактирование ячеек с заливкой
        """
        self.path = filename
        self.work_book = load_workbook(filename)
        self.readonly_fill_color = readonly_fill_color

    def save(self, period: Period):
        """Сохранение обработанного файла в базу данных."""
        extract_sheets: list[BuildSheet] = self.extract()
        for position, extract_sheet in enumerate(extract_sheets):
            sheet = Sheet.objects.create(
                name=extract_sheet.name,
                position=position,
                period=period
            )
            extract_sheet.cache_container.save(sheet.pk)
            # Соотношение позиции и созданных идентификаторов
            columns_mapper: dict[int, int] = {}
            rows_mapper: dict[int, int] = {}
            columns_styles: dict[int, dict] = {}
            rows_styles: dict[int, dict] = {}
            for cell in extract_sheet.cells:
                column_id: int = cell.column_id
                row_id: int = cell.row_id
                columns_styles[column_id] = {}
                rows_styles[row_id] = {}

                if column_id not in columns_mapper:
                    column_parameters = {
                        'index': column_id,
                        'fixed': extract_sheet.fixed_column is not None and column_id < extract_sheet.fixed_column
                    }
                    if column_id in extract_sheet.columns_dimension:
                        cd: dict = asdict(extract_sheet.columns_dimension[column_id])
                        columns_styles[column_id] = cd.pop('style')
                        column_parameters = {**column_parameters, **cd}
                    column: ColumnDimension = ColumnDimension.objects.create(sheet=sheet, **column_parameters)
                    columns_mapper[column_id] = column.id

                if row_id not in rows_mapper:
                    row_parameters = {
                        'index': row_id,
                        'fixed': extract_sheet.fixed_row is not None and row_id < extract_sheet.fixed_row
                    }
                    if row_id in extract_sheet.rows_dimension:
                        rd = asdict(extract_sheet.rows_dimension[row_id])
                        rows_styles[row_id] = rd.pop('style')
                        row_parameters = {**row_parameters, **rd}
                    row: RowDimension = RowDimension.objects.create(sheet=sheet, **row_parameters)
                    rows_mapper[row_id] = row.id

                cell.column_id = columns_mapper[column_id]
                cell.row_id = rows_mapper[row_id]
                # Объединяем стили cell <- row <- col
                Cell.objects.create(**{
                    **columns_styles[column_id],
                    **rows_styles[row_id],
                    **{k: v for k, v in asdict(cell).items() if k != 'coordinate'}
                })

            for merged_cell in extract_sheet.merged_cells:
                MergedCell.objects.create(sheet=sheet, **asdict(merged_cell))

    def extract(self) -> list[BuildSheet]:
        """Парсинг файла Excel.

        Функция создает структуру данных, которая является первоначальной обработкой.
        После выделения необходимых данных можно осуществлять транзакционную запись в базу данных.
        Структура данных может использоваться для предварительной демонстрации планируемого отчета.
        """
        sheets: list[BuildSheet] = []
        for sheet in self.work_book.worksheets:
            name: str = sheet.title
            columns_dimension: dict[int, BuildColumnDimension] = self._parse_columns_dimension(sheet.column_dimensions)
            rows_dimension: dict[int, BuildRowDimension] = self._parse_rows_dimension(sheet.row_dimensions)
            cells: list[BuildCell] = self._parse_cells(self.work_book, sheet.rows)
            merged_cells: list[BuildMergedCell] = self._parse_merged_cells(sheet.merged_cells.ranges)
            [fixed_column_name, fixed_row_index] = coordinate_from_string(
                sheet.freeze_panes
            ) if sheet.freeze_panes is not None else (None, None)
            sheets.append(BuildSheet(
                name=name,
                columns_dimension=columns_dimension,
                fixed_column=column_index_from_string(fixed_column_name) if fixed_column_name else None,
                rows_dimension=rows_dimension,
                fixed_row=fixed_row_index,
                cells=cells,
                merged_cells=merged_cells
            ))
        return self.evaluate_cells(sheets)

    def _parse_columns_dimension(self, holder: DimensionHolder) -> dict[int, BuildColumnDimension]:
        """Парсинг имеющихся колонок."""
        return {
            column_index_from_string(col_letter): BuildColumnDimension(
                column.width * 7,
                column.hidden,
                self.__border_style(column)
            ) for col_letter, column in holder.items()
        }

    def _parse_rows_dimension(self, holder: DimensionHolder) -> dict[int, BuildRowDimension]:
        """Парсинг имеющихся строк."""
        return {
            row.index: BuildRowDimension(int(row.height * 1.2) if row.height else None, self.__border_style(row))
            for index, row in holder.items()
        }

    @staticmethod
    def __border_style(dimension: OpenpyxlRowDimension | OpenpyxlColumnDimension | Cell | MergedCell) -> BuildStyle:
        """for p in ('top', 'bottom', 'left', 'right', 'diagonal', 'diagonalDown', 'diagonalUp',)"""
        return BuildStyle(
            dimension.alignment.horizontal,
            dimension.alignment.vertical,
            dimension.font.sz,
            dimension.font.b,
            dimension.font.i,
            dimension.font.strike or False,
            dimension.font.u,
            f'#{dimension.font.color.value[2:]}'
            if dimension.font.color and dimension.font.color.type == 'rgb' else '#000000',
            '#FFFFFF' if dimension.fill.patternType is None else f'#{dimension.fill.fgColor.value[2:]}',
            {
                p: getattr(dimension.border, p).style
                for p in ('top', 'bottom', 'left', 'right', 'diagonal')
            }
        )

    @staticmethod
    def __color_transform(wb: Workbook, color: Color):
        try:
            if color and color.type == 'indexed':
                if color.index == 64 or color.index == 65:
                    color = None
                else:
                    color.value = COLOR_INDEX[color.index]
            elif color and color.type == 'theme':
                color.type = 'rgb'
                color.value = theme_and_tint_to_rgb(wb, color.theme, color.tint)
            return color
        except TypeError:
            return None

    def _parse_cells(self, wb: Workbook, rows: Iterator[tuple[OpenpyxlCell | OpenpyxlMergedCell]]) -> list[BuildCell]:
        """Парсинг ячеек.

        Переданный параметр rows представляет собой матрицу.
        Каждая строка включает в себя массив ячеек, который соотноситься с колонками.
        """
        cells: list[BuildCell] = []
        for row in rows:
            for cell in row:
                fill_color = self.__color_transform(wb, cell.fill.fgColor)
                font_color = self.__color_transform(wb, cell.font.color)

                # Временная заглушка
                if (font_color and font_color.index == 1 and cell.fill.patternType is None) or \
                        (font_color and font_color.index == 1 and fill_color.value == WHITE):
                    font_color.type = 'rgb'
                    font_color.value = '00000000'
                kind, number_format = self._get_cell_kind_and_number_format(cell)
                cells.append(BuildCell(
                    column_id=cell.column,
                    row_id=cell.row,
                    kind=kind,
                    editable=not self.readonly_fill_color or fill_color.value == '00000000',
                    coordinate=cell.coordinate,
                    formula=self._get_cell_formula(cell),
                    number_format=number_format,
                    comment=cell.comment,
                    default=self._get_cell_default(cell),
                    border_color=self._get_cell_border_color(wb, cell),
                    **asdict(self.__border_style(cell))
                ))
        return cells

    @staticmethod
    def _parse_merged_cells(ranges: list[OpenpyxlMergedCell]) -> list[BuildMergedCell]:
        """Парсинг объединенных ячеек."""
        return [BuildMergedCell(rng.min_col, rng.min_row, rng.max_col, rng.max_row) for rng in ranges]

    def evaluate_cells(self, sheets: list[BuildSheet]) -> list[BuildSheet]:
        """Предварительно рассчитываем значения ячеек.

        Excel не хранит кешированные значения, вместо этого он хранит формулы.
        Нам необходимо рассчитать формулы, однако значения могут быть перекрестными.
        Поэтому нам необходимо собирать единую структуру и каждый раз формировать модель.
        """

        cells_values: dict[str, str] = {}
        for sheet in sheets:
            cells_values.update({
                self.coordinate(sheet.name, cell.column_id, cell.row_id): cell.default or 0 for cell in sheet.cells
            })
        for sheet in sheets:
            evaluate_model = ModelCompiler().read_and_parse_dict(cells_values, default_sheet=sheet.name)
            evaluator = Evaluator(evaluate_model)
            for cell in sheet.cells:
                if cell.formula:
                    coordinate = self.coordinate(sheet.name, cell.column_id, cell.row_id)
                    success, value = evaluate_formula(evaluator, coordinate)
                    cell.default = value if success else None
                    cell.default_error = value if not success else None
                    # cells_values - для локального пересчета, если все ок - изменяем, если нет, ничего не делаем
                    if success:
                        cells_values[coordinate] = cell.default
                    sheet.cache_container.add_formula(cell.coordinate, cell.formula)
        return sheets

    @staticmethod
    def coordinate(sheet: str, column: int, row: int) -> str:
        """Получаем координату."""
        return f'{sheet}!{get_column_letter(column)}{row}'

    def _get_cell_kind_and_number_format(self, cell: OpenpyxlCell) -> tuple[str, str | None]:
        """Получение типа и форматирования чисел для ячейки."""
        if cell.data_type in (KindCell.NUMERIC, KindCell.DATE, TYPE_FORMULA):
            k = self._NUMBER_FORMAT_KIND_MAP.get(cell.number_format)
            data_type = KindCell.NUMERIC if cell.data_type == TYPE_FORMULA else cell.data_type
            return k or (data_type, cell.number_format)
        if '\n' in cell.value:
            return KindCell.TEXT, None
        return cell.data_type, None

    @staticmethod
    def _get_cell_formula(cell: OpenpyxlCell) -> str | None:
        """Получение формулы для ячейки."""
        if isinstance(cell.value, str) and cell.value and cell.value[0] == '=':
            return cell.value
        return None

    @staticmethod
    def _get_cell_default(cell: OpenpyxlCell) -> str | None:
        """Получение значения по умолчанию для ячейки."""
        if cell.value is not None:
            return str(cell.value)
        return None

    def _get_cell_border_color(self, wb: Workbook, cell: OpenpyxlCell) -> dict[str, str]:
        """Получение цвета границы для ячейки."""
        border_color = {
            positional: self.__color_transform(wb, getattr(cell.border, positional).color)
            for positional in ('top', 'bottom', 'left', 'right', 'diagonal')
        }
        return {
            p: f'#{c.value[2:]}' if c and c.type == 'rgb' else None
            for p, c in border_color.items()
        }

    _NUMBER_FORMAT_KIND_MAP = {
        'General': (KindCell.NUMERIC, '0.00'),
        '@': (KindCell.STRING, None),
        'mm-dd-yy': (KindCell.DATE, 'dd.mm.yyyy'),
        '[$-F800]dddd\,\ mmmm\ dd\,\ yyyy': (KindCell.DATE, 'dd.mm.yyyy'),
        '[$-F400]h:mm:ss\ AM/PM': (KindCell.TIME, 'hh:mm:ss'),
    }
