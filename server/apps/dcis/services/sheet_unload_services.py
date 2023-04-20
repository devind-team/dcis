"""Модуль для выгрузки листов."""

import operator
from abc import ABC, abstractmethod
from dataclasses import dataclass
from enum import Enum, auto
from functools import reduce
from statistics import mean
from typing import Any, Iterable, Sequence

from django.db.models import Model, Q, QuerySet
from openpyxl.utils import column_index_from_string, get_column_letter

from apps.dcis.models import Document
from apps.dcis.models.sheet import Cell, ColumnDimension, KindCell, MergedCell, RowDimension, Sheet, Value
from apps.dcis.permissions import (
    AddChildRowDimensionBase,
    ChangeChildRowDimensionHeightBase,
    ChangeDocumentSheetBase,
    ChangeValueBase,
    DeleteChildRowDimensionBase,
)


class DataUnloader(ABC):
    """Выгрузчик данных."""

    def __init__(self) -> None:
        self.data: list[dict] | dict | None = None

    @abstractmethod
    def unload_data(self) -> list[dict] | dict:
        """Выгрузка данных."""
        ...

    def unload(self) -> list[dict] | dict:
        """Выгрузка данных с учетом кеша."""
        if self.data is not None:
            return self.data
        self.data = self.unload_data()
        return self.data

    @classmethod
    def unload_raw_data(
        cls,
        objects: QuerySet | Sequence[Model],
        fields: Sequence[str],
        properties: Sequence[str] | None = None
    ) -> list[dict]:
        """Выгрузка необработанных данных."""
        if properties is not None:
            result: list[dict] = []
            for obj in (objects.all() if isinstance(objects, QuerySet) else objects):
                obj_dict = cls._model_to_dict(obj, fields=fields)
                for pr in properties:
                    obj_dict[pr] = getattr(obj, pr)
                result.append(obj_dict)
            return result
        if isinstance(objects, QuerySet):
            return list(objects.values(*fields))
        return [cls._model_to_dict(obj, fields=fields) for obj in objects]

    @staticmethod
    def _model_to_dict(instance: Model, fields: Sequence[str] | None = None) -> dict:
        """Преобразование модели в словарь."""
        return {k: v for k, v in instance.__dict__.items() if k in fields}


class SheetColumnsUnloader(DataUnloader):
    """Выгрузчик колонок листа."""

    def __init__(self, columns: QuerySet[ColumnDimension] | Sequence[ColumnDimension]) -> None:
        super().__init__()
        self.columns = columns

    def unload_data(self) -> list[dict] | dict:
        """Выгрузка колонок листа."""
        columns = self.unload_raw_data(self.columns, self._column_fields)
        for column in columns:
            column['name'] = get_column_letter(column['index'])
        return columns

    _column_fields = (
        'id', 'index', 'width',
        'fixed', 'hidden', 'kind',
        'created_at', 'updated_at', 'user_id'
    )


class SheetRowsUnloaderBase(DataUnloader, ABC):
    """Базовый выгрузчик строк листа."""

    def __init__(
        self,
        columns_unloader: SheetColumnsUnloader,
        rows: QuerySet[RowDimension] | Sequence[RowDimension],
        cells: QuerySet[Cell] | Sequence[Cell],
        merged_cells: QuerySet[MergedCell] | Sequence[MergedCell],
        values: QuerySet[Value] | Sequence[Value]
    ):
        super().__init__()
        self.columns_unloader = columns_unloader
        self.rows = rows
        self.cells = cells
        self.merged_cells = merged_cells
        self.values = values

    _rows_fields = (
        'id', 'index', 'height',
        'fixed', 'hidden', 'dynamic',
        'aggregation', 'created_at', 'updated_at',
        'parent_id', 'document_id', 'object_id',
        'user_id',
    )
    _cells_fields = (
        'id', 'kind', 'editable',
        'formula', 'number_format', 'comment',
        'default', 'default_error', 'mask',
        'tooltip', 'column_id', 'row_id',
        'horizontal_align', 'vertical_align', 'size',
        'strong', 'italic', 'strike',
        'underline', 'color', 'background',
        'border_style', 'border_color', 'aggregation',
    )
    _values_fields = (
        'document_id', 'column_id', 'row_id', 'value', 'error',
    )
    _merged_cells_fields = (
        'min_col', 'max_col',
    )
    _merged_cells_properties = (
        'colspan', 'rowspan', 'target', 'cells',
    )

    def _unload_raw_rows(self) -> list[dict]:
        """Выгрузка необработанных строк листа."""
        return self.unload_raw_data(self.rows, self._rows_fields)

    def _unload_raw_cells(self) -> list[dict]:
        """Выгрузка необработанных ячеек листа."""
        return self.unload_raw_data(self.cells, self._cells_fields)

    def _unload_raw_values(self) -> list[dict]:
        """Выгрузка необработанных значений листа."""
        return self.unload_raw_data(self.values, self._values_fields)

    def _unload_raw_merged_cells(self) -> list[dict]:
        """Выгрузка необработанных объединенных ячеек."""
        return self.unload_raw_data(self.merged_cells, self._merged_cells_fields, self._merged_cells_properties)

    @staticmethod
    def _create_columns_map(columns: list[dict]) -> dict[dict]:
        """Создание структуры для быстрого поиска колонок по id."""
        return reduce(lambda a, c: {**a, c['id']: c}, columns, {})

    @staticmethod
    def _create_merged_cells_map(merged_cells: list[dict]) -> dict[dict]:
        """Создание структуры для быстрого поиска объединенных ячеек по позиции основной ячейки."""
        return reduce(lambda a, c: {**a, c['target']: c}, merged_cells, {})

    @staticmethod
    def _create_merged_cell_positions(merged_cells: list[dict]) -> list[str]:
        """Создание списка позиций объединенных ячеек."""
        return reduce(lambda a, c: [*a, *c['cells']], merged_cells, [])

    @staticmethod
    def _create_merged_cell_row_positions(merged_cells: list[dict]) -> list[str]:
        """Создание списка позиций объединенных ячеек для первой строки."""
        return reduce(lambda a, c: [*a, *c['cells'][:c['max_col'] - c['min_col'] + 1]], merged_cells, [])

    def _prepare_data(self, rows: list[dict], cells: list[dict]) -> None:
        """Подготовка строк и ячеек к выгрузке после сортировки строк."""
        columns = self.columns_unloader.unload()
        merged_cells = self._unload_raw_merged_cells()
        values = self._unload_raw_values()
        columns_map = self._create_columns_map(columns)
        merged_cells_map = self._create_merged_cells_map(merged_cells)
        merged_cell_positions = self._create_merged_cell_positions(merged_cells)
        merged_cell_row_positions = self._create_merged_cell_row_positions(merged_cells)
        self._add_cells(rows, columns_map, cells)
        self._add_cell_values(rows, values)
        self._add_cell_properties(rows, columns_map, merged_cells_map)
        self._filter_rows_cells(rows, merged_cells_map, merged_cell_positions, merged_cell_row_positions)

    @classmethod
    def _connect_rows(cls, rows: list[dict]) -> list[dict]:
        """Создание деревьев строк."""
        for row in rows:
            row['parent'] = None
            row['children'] = []
        trees = [row for row in rows if row['parent_id'] is None]
        for root in trees:
            cls._add_children(rows, root)
        return trees

    @classmethod
    def _add_children(cls, rows: list[dict], row: dict) -> None:
        """Добавление дочерних строк к строке."""
        row['children'] = [r for r in rows if r['parent_id'] == row['id']]
        for child in row['children']:
            child['parent'] = row
            cls._add_children(rows, child)

    def _flatten_rows(self, rows_tree: list[dict], sort: bool = True) -> list[dict]:
        """Превращение строк в плоскую структуру."""
        result: list[dict] = []
        if sort:
            rows_tree = self._sort_rows(rows_tree)
        for row in rows_tree:
            result.append(row)
            result.extend(self._flatten_rows(row['children']))
        return result

    @abstractmethod
    def _sort_rows(self, rows: list[dict]) -> list[dict]:
        """Сортировка строк."""
        ...

    @staticmethod
    @abstractmethod
    def _add_cells(rows: list[dict], columns_map: dict[dict], cells: list[dict]) -> None:
        """Добавление ячеек к строкам."""
        ...

    @staticmethod
    @abstractmethod
    def _add_cell_values(rows: list[dict], values: list[dict]) -> None:
        """Добавление значений к ячейкам."""
        ...

    @classmethod
    def _add_cell_properties(cls, rows: list[dict], columns_map: dict[dict], merged_cells_map: dict[dict]) -> None:
        """Добавление свойств к ячейкам."""
        for row in rows:
            for cell in row['cells']:
                column = columns_map[cell['column_id']]
                cls._add_cell_positions(row, column, cell)
                cls._add_cell_spans(row, cell, merged_cells_map)
                cls._add_cell_related_positions(row, column, cell)

    @staticmethod
    def _add_cell_positions(row: dict, column: dict, cell: dict) -> None:
        """Добавление позиций для ячейки."""
        cell['position'] = f'{column["name"]}{row["name"]}'
        cell['global_position'] = f'{column["name"]}{row["global_index"]}'

    @classmethod
    def _add_cell_spans(cls, row: dict, cell: dict, merged_cells_map: dict) -> None:
        """Добавление объединений по колонкам и строкам для ячейки."""
        root_cell = cls._find_root_cell(row, cell)
        merged_cell = merged_cells_map.get(root_cell['position'], None)
        cell['colspan'] = merged_cell['colspan'] if merged_cell else 1
        cell['rowspan'] = merged_cell['rowspan'] if merged_cell and cell['id'] == root_cell['id'] else 1

    @staticmethod
    def _add_cell_related_positions(row: dict, column: dict, cell: dict):
        """Добавление связанных с объединением позиций."""
        cell['related_global_positions'] = []
        for row_offset in range(cell['rowspan']):
            for column_offset in range(cell['colspan']):
                column_name = get_column_letter(column_index_from_string(column['name']) + column_offset)
                row_index = row['global_index'] + row_offset
                cell['related_global_positions'].append(f'{column_name}{row_index}')

    @classmethod
    def _find_root_cell(cls, row: dict, cell: dict) -> dict:
        """Нахождение корневой ячейки для ячейки дочерней строки.
        Корневой считается ячейка лежащая в той же колонке в строке верхнего уровня.
        """
        if row['parent'] is None:
            return cell
        return cls._find_root_cell(
            row['parent'],
            next(
                parent_cell for parent_cell in row['parent']['cells']
                if parent_cell['column_id'] == cell['column_id']
            )
        )

    @classmethod
    def _filter_rows_cells(
        cls,
        rows: list[dict],
        merged_cells_map: dict,
        merged_cell_positions: list[str],
        merged_cell_row_positions: list[str]
    ) -> None:
        """Удаление лишних ячеек из строк."""
        for row in rows:
            row['output_cells']: list[dict] = []
            for cell in row['cells']:
                root_cell = cls._find_root_cell(row, cell)
                positions = merged_cell_positions if cell['id'] == root_cell['id'] else merged_cell_row_positions
                if root_cell['position'] in merged_cells_map or root_cell['position'] not in positions:
                    row['output_cells'].append(cell)
        for row in rows:
            row['cells'] = row['output_cells']


class SheetRowsUnloader(SheetRowsUnloaderBase):
    """Выгрузчик строк листа."""

    def __init__(
        self,
        columns_unloader: SheetColumnsUnloader,
        rows: QuerySet[RowDimension] | Sequence[RowDimension],
        cells: QuerySet[Cell] | Sequence[Cell],
        merged_cells: QuerySet[MergedCell] | Sequence[MergedCell],
        values: QuerySet[Value] | Sequence[Value]
    ) -> None:
        super().__init__(columns_unloader, rows, cells, merged_cells, values)

    def unload_data(self) -> list[dict] | dict:
        """Выгрузка строк листа."""
        rows = self._unload_raw_rows()
        cells = self._unload_raw_cells()
        row_trees = self._connect_rows(rows)
        self._add_row_names(row_trees)
        rows = self._flatten_rows(row_trees)
        self._add_global_indices(rows)
        self._prepare_data(rows, cells)
        return rows

    def _sort_rows(self, rows: list[dict]) -> list[dict]:
        """Сортировка строк."""
        return sorted(rows, key=lambda r: r['index'])

    @staticmethod
    def _add_cells(rows: list[dict], columns_map: dict[dict], cells: list[dict]) -> None:
        """Добавление ячеек к строкам."""
        for row in rows:
            row['cells'] = [cell for cell in cells if cell['row_id'] == row['id']]
            row['cells'].sort(key=lambda cell: columns_map[cell['column_id']]['index'])

    @staticmethod
    def _add_cell_values(rows: list[dict], values: list[dict]) -> None:
        """Добавление значений к ячейкам."""
        for row in rows:
            for cell in row['cells']:
                val = cell['default']
                del cell['default']
                err = cell['default_error']
                del cell['default_error']
                value = next(
                    (v for v in values if v['row_id'] == cell['row_id'] and v['column_id'] == cell['column_id']),
                    None
                )
                if value is not None:
                    val = value['value']
                    err = value['error']
                cell.update({'value': val, 'error': err})

    @classmethod
    def _add_row_names(cls, rows_tree: list[dict]) -> None:
        """Добавление имен к строкам."""
        for row in rows_tree:
            row['name'] = cls._get_row_name(row)
            if len(row['children']):
                cls._add_row_names(row['children'])

    @classmethod
    def _get_row_name(cls, row: dict, indices: list[int] | None = None) -> str:
        """Получение имени строки."""
        indices = indices or []
        if row['parent'] is not None:
            return cls._get_row_name(row['parent'], [str(row['index']), *indices])
        return '.'.join([str(row['index']), *indices])

    @staticmethod
    def _add_global_indices(rows: list[dict]) -> None:
        """Добавление индексов в плоской структуре для строк."""
        for i, row in enumerate(rows, 1):
            row['global_index'] = i


class SheetPartialRowsUploader(SheetRowsUnloader):
    """Выгрузчик подмножества строк листа."""

    def __init__(
        self,
        columns_unloader: SheetColumnsUnloader,
        rows: QuerySet[RowDimension] | Sequence[RowDimension],
        cells: QuerySet[Cell] | Sequence[Cell],
        merged_cells: QuerySet[MergedCell] | Sequence[MergedCell],
        values: QuerySet[Value] | Sequence[Value],
        rows_global_indices_map: dict[dict]
    ) -> None:
        super().__init__(
            columns_unloader=columns_unloader,
            rows=rows,
            cells=cells,
            merged_cells=merged_cells,
            values=values
        )
        self.rows_global_indices_map = rows_global_indices_map

    def unload_data(self) -> list[dict] | dict:
        """Частичная выгрузка строк листа."""
        rows = self._unload_raw_rows()
        cells = self._unload_raw_cells()
        parent_rows = self._find_row_parents(rows)
        row_trees = self._connect_rows([*rows, *parent_rows])
        self._add_row_names(row_trees)
        all_rows = self._flatten_rows(row_trees)
        all_cells = [*cells, *self._find_rows_cells(parent_rows)]
        self._add_global_indices(all_rows)
        self._prepare_data(all_rows, all_cells)
        return self._sort_rows_finally(rows)

    @classmethod
    def _find_row_parents(cls, rows: list[dict]) -> list[dict]:
        """Поиск родительских строк."""
        parent_rows = cls.unload_raw_data(
            RowDimension.objects.filter(pk__in=[row['parent_id'] for row in rows if row['parent_id'] is not None]),
            cls._rows_fields
        )
        unique_parent_rows = [
            parent_row for parent_row in parent_rows if
            next((row for row in rows if row['id'] == parent_row['id']), None) is None
        ]
        if len(unique_parent_rows):
            return [*unique_parent_rows, *cls._find_row_parents(unique_parent_rows)]
        return unique_parent_rows

    @classmethod
    def _find_rows_cells(cls, rows: list[dict]) -> list[dict]:
        """Поиск ячеек для строк."""
        return cls.unload_raw_data(Cell.objects.filter(row__id__in=[row['id'] for row in rows]), cls._cells_fields)

    def _add_global_indices(self, rows: list[dict]) -> None:
        """Добавление индексов в плоской структуре для строк."""
        for row in rows:
            row['global_index'] = self.rows_global_indices_map[row['id']]

    @staticmethod
    def _sort_rows_finally(rows: list[dict]) -> list[dict]:
        """Сортировка строк."""
        return sorted(rows, key=lambda row: row['global_index'])


@dataclass
class ReportDocument:
    """Документ для выгрузки сводного отчета."""
    document: Document
    is_visible: bool
    color: str


class ReportAggregation(Enum):
    """Тип агрегации для сводного отчета."""
    CONCAT = auto()
    SUM = auto()
    AVG = auto()
    MIN = auto()
    MAX = auto()

    def is_applicable(self, kind: str) -> bool:
        """Применима ли агрегация к типу."""
        return kind in _AGGREGATION_TYPE_MAP[self]

    def apply(self, values: Iterable[str]) -> str:
        """Применение агрегации."""
        return getattr(self, self.name.lower())(values)

    @staticmethod
    def concat(values: Iterable[str]) -> str:
        """Применение функции CONCAT."""
        return reduce(operator.add, values, '')

    @staticmethod
    def sum(values: Iterable[str]) -> str:
        """Применение функции SUM."""
        return str(sum(float(v) for v in values))

    @staticmethod
    def avg(values: Iterable[str]) -> str:
        """Применение функции AVG."""
        return str(mean(float(v) for v in values))

    @staticmethod
    def min(values: Iterable[str]) -> str:
        """Применение функции MIN."""
        return str(min(float(v) for v in values))

    @staticmethod
    def max(values: Iterable[str]) -> str:
        """Применение функции MAX."""
        return str(max(float(v) for v in values))


_AGGREGATION_TYPE_MAP = {
    ReportAggregation.CONCAT: [KindCell.STRING, KindCell.TEXT],
    ReportAggregation.SUM: [KindCell.NUMERIC],
    ReportAggregation.AVG: [KindCell.NUMERIC],
    ReportAggregation.MIN: [KindCell.NUMERIC],
    ReportAggregation.MAX: [KindCell.NUMERIC],
}


class SheetReportRowsUnloader(SheetRowsUnloaderBase):
    """Выгрузчик строк листа для сводного отчета."""

    def __init__(
        self,
        columns_unloader: SheetColumnsUnloader,
        rows: QuerySet[RowDimension] | Sequence[RowDimension],
        cells: QuerySet[Cell] | Sequence[Cell],
        merged_cells: QuerySet[MergedCell] | Sequence[MergedCell],
        values: QuerySet[Value] | Sequence[Value],
        report_documents: list[ReportDocument],
        main_document: Document | None,
        aggregation: ReportAggregation | None,
        expanded_row_groups: dict[int, bool],
        indices_groups: list[list[int]]
    ) -> None:
        super().__init__(columns_unloader, rows, cells, merged_cells, values)
        self.report_documents = report_documents
        self.main_document = main_document
        self.aggregation = aggregation
        self.expanded_row_groups = expanded_row_groups
        self.indices_groups = indices_groups

    def unload_data(self) -> list[dict] | dict:
        """Выгрузка строк листа для сводного отчета."""
        rows = self._unload_raw_rows()
        cells = self._unload_raw_cells()
        row_trees = self._connect_rows(rows)
        row_trees = self._expand_rows(row_trees)
        rows = self._flatten_rows(row_trees, False)
        self._add_rows_additional_data(rows)
        self._prepare_data(rows, cells)
        return rows

    def _prepare_data(self, rows: list[dict], cells: list[dict]) -> None:
        """Подготовка строк и ячеек к выгрузке после сортировки строк."""
        super()._prepare_data(rows, cells)
        self._set_ids(rows)

    def _sort_rows(self, rows: list[dict]) -> list[dict]:
        """Сортировка строк по документу, а затем по индексу."""
        return sorted(
            rows,
            key=lambda r: (
                next(
                    (rd[0] for rd in enumerate(self.report_documents, 1) if r['document_id'] == rd[1].document.id), 0
                ),
                r['index']
            )
        )

    @staticmethod
    def _add_cells(rows: list[dict], columns_map: dict[dict], cells: list[dict]) -> None:
        """Добавление ячеек к строкам."""
        for row in rows:
            row['cells'] = [{**cell} for cell in cells if cell['row_id'] == row['id']]
            row['cells'].sort(key=lambda cell: columns_map[cell['column_id']]['index'])

    def _add_cell_values(self, rows: list[dict], values: list[dict]) -> None:
        """Добавление значений к ячейкам."""
        for row in rows:
            for cell in row['cells']:
                val = cell['default']
                del cell['default']
                err = cell['default_error']
                del cell['default_error']
                cell_values = [
                    v for v in values if
                    v['row_id'] == cell['row_id'] and v['column_id'] == cell['column_id']
                ]
                if row['parent_id'] is None:
                    if row['document_id'] is not None:
                        value = next((v for v in cell_values if v['document_id'] == row['document_id']), None)
                        if value is not None:
                            val = value['value']
                            err = value['error']
                    elif self._is_aggregation_applicable(cell, cell_values):
                        val = self.aggregation.apply(v['value'] for v in cell_values)
                        err = None
                    elif self.main_document is not None:
                        value = next((v for v in cell_values if v['document_id'] == self.main_document.id), None)
                        if value is not None:
                            val = value['value']
                            err = value['error']
                else:
                    val = cell_values[0]['value']
                    err = cell_values[0]['error']
                cell.update({'value': val, 'error': err})

    def _expand_rows(self, rows: list[dict]) -> list[dict]:
        """Расширение строк.

        Расширение заключается в дублировании групп строк с общими ячейками для каждого документа.
        Подстроки разных документов становятся подстроками соответствующей новой строки.
        Необходимость дублировать ту или иную строку определяет словарь `expanded_row_groups`."""
        expended_rows: list[dict] = []
        for i, group in enumerate(self.indices_groups):
            if self.expanded_row_groups[i]:
                for document_id in (rd.document.id for rd in self.report_documents):
                    for index in group:
                        row = next(r for r in rows if r['index'] == index)
                        cloned_row = {
                            **row,
                            'children': [child for child in row['children'] if child['document_id'] == document_id],
                            'document_id': document_id,
                            'is_cloned': True
                        }
                        for child in cloned_row['children']:
                            child['parent'] = cloned_row
                        expended_rows.append(cloned_row)
            else:
                for index in group:
                    expended_rows.append(next(r for r in rows if r['index'] == index))
        return expended_rows

    def _add_rows_additional_data(self, rows: list[dict]) -> None:
        """Добавление дополнительных данных для строк."""
        colors: dict[int, str] = {}
        for rd in self.report_documents:
            colors[rd.document.id] = rd.color
        for i, row in enumerate(rows, 1):
            row['global_index'] = i
            row['name'] = str(i)
            if row['document_id'] is not None:
                row['background'] = colors[row['document_id']]

    def _is_aggregation_applicable(self, cell: dict, values: list[dict]) -> bool:
        """Применима ли агрегация к ячейке и значениям."""
        return (
            self.aggregation is not None and
            self.aggregation.is_applicable(cell['kind']) and
            len(values) > 1 and
            all(v['error'] is None for v in values)
        )

    @staticmethod
    def _set_ids(rows: list[dict]) -> None:
        """Установка идентификаторов для тех строк и ячеек, которые были клонированы."""
        for row in rows:
            if 'is_cloned' in row and row['is_cloned']:
                row['id'] = f'{row["id"]} {row["document_id"]}'
                for cell in row['cells']:
                    cell['id'] = f'{cell["id"]} {row["document_id"]}'
                    cell['row_id'] = row['id']


class SheetUnloader(DataUnloader):
    """Выгрузчик листа."""

    def __init__(self, sheet: Sheet, fields: Sequence[str]) -> None:
        super().__init__()
        self.sheet = sheet
        self.fields = fields
        self.columns_unloader: SheetColumnsUnloader | None = None

    @abstractmethod
    def unload_rows(self) -> list[dict] | dict:
        """Выгрузка строк."""
        ...

    @abstractmethod
    def get_permissions(self) -> dict[str, bool]:
        """Получение разрешений."""
        ...

    def unload_data(self) -> list[dict] | dict:
        """Выгрузка листа."""
        sheet = self._model_to_dict(
            self.sheet,
            fields=[field for field in self.fields if field not in ('columns', 'rows', 'can_change')]
        )
        self.columns_unloader = SheetColumnsUnloader(self.sheet.columndimension_set.all())
        if 'columns' in self.fields:
            sheet['columns'] = self.columns_unloader.unload()
        if 'rows' in self.fields:
            sheet['rows'] = self.unload_rows()
        if 'can_change' in self.fields:
            sheet.update(self.get_permissions())
        return sheet


class PeriodSheetUnloader(SheetUnloader):
    """Выгрузчик листа для периода."""

    def unload_rows(self) -> list[dict] | dict:
        """Выгрузка строк."""
        rows = self.sheet.rowdimension_set.filter(parent__isnull=True)
        return SheetRowsUnloader(
            columns_unloader=self.columns_unloader,
            rows=rows,
            cells=Cell.objects.filter(row__in=[row.id for row in rows]),
            merged_cells=self.sheet.mergedcell_set.all(),
            values=Value.objects.none(),
        ).unload()

    def get_permissions(self) -> dict[str, bool]:
        return {
            'can_change': True,
            'can_change_value': True,
            'can_add_child_row_dimension': True,
            'can_change_child_row_dimension_height': True,
            'can_delete_child_row_dimension': True,
        }


class DocumentSheetUnloader(SheetUnloader):
    """Выгрузчик листа с документом."""

    def __init__(self, context: Any, sheet: Sheet, document_id: int | str, fields: Sequence[str]) -> None:
        super().__init__(sheet, fields)
        self.document = Document.objects.get(pk=document_id)
        self.change_document_sheet = ChangeDocumentSheetBase(context.user, self.document)
        self.change_value = ChangeValueBase(context.user, self.document)
        self.add_child_row_dimension = AddChildRowDimensionBase(context.user, self.document)
        self.change_child_row_dimension_height = ChangeChildRowDimensionHeightBase(context.user, self.document)
        self.delete_child_row_dimension = DeleteChildRowDimensionBase(context.user, self.document)

    def unload_rows(self) -> list[dict] | dict:
        """Выгрузка строк."""
        rows = self.sheet.rowdimension_set.filter(
            Q(parent__isnull=True) | Q(parent__isnull=False, document_id=self.document.id)
        )
        return SheetRowsUnloader(
            columns_unloader=self.columns_unloader,
            rows=rows,
            cells=Cell.objects.filter(row__in=[row.id for row in rows]),
            merged_cells=self.sheet.mergedcell_set.all(),
            values=self.sheet.value_set.filter(document_id=self.document.id)
        ).unload()

    def get_permissions(self) -> dict[str, bool]:
        return {
            'can_change': self.change_document_sheet.has_permission,
            'can_change_value': self.change_value.has_privilege,
            'can_add_child_row_dimension': self.add_child_row_dimension.has_privilege,
            'can_change_child_row_dimension_height': self.change_child_row_dimension_height.has_privilege,
            'can_delete_child_row_dimension': self.delete_child_row_dimension.has_privilege,
        }


class ReportSheetUnloader(SheetUnloader):
    """Выгрузчик листа для сводного отчета."""

    def __init__(
        self,
        sheet: Sheet,
        report_documents: list[ReportDocument],
        main_document: Document | None,
        aggregation: ReportAggregation | None,
        expanded_row_groups: dict[int, bool],
        fields: Sequence[str]
    ) -> None:
        super().__init__(sheet, fields)
        self.report_documents = report_documents
        self.main_document = main_document
        self.aggregation = aggregation
        self.expanded_row_groups = expanded_row_groups

    def unload_rows(self) -> list[dict] | dict:
        """Выгрузка строк."""
        from apps.dcis.services.row_dimension_services import get_indices_groups_to_expand

        all_documents = [dr.document for dr in self.report_documents]
        visible_documents = [dr.document for dr in self.report_documents if dr.is_visible]
        rows = self.sheet.rowdimension_set.filter(
            Q(parent__isnull=True) | Q(parent__isnull=False, document__in=visible_documents)
        )
        return SheetReportRowsUnloader(
            columns_unloader=self.columns_unloader,
            rows=rows,
            cells=Cell.objects.filter(row__in=[row.id for row in rows]),
            merged_cells=self.sheet.mergedcell_set.all(),
            values=Value.objects.filter(row__in=rows, document__in=all_documents),
            report_documents=self.report_documents,
            main_document=self.main_document,
            aggregation=self.aggregation,
            expanded_row_groups=self.expanded_row_groups,
            indices_groups=get_indices_groups_to_expand(self.sheet),
        ).unload()

    def get_permissions(self) -> dict[str, bool]:
        return {
            'can_change': False,
            'can_change_value': False,
            'can_add_child_row_dimension': False,
            'can_change_child_row_dimension_height': False,
            'can_delete_child_row_dimension': False,
        }
