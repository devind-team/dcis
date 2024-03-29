"""Сервис для загрузки данных их файла."""
import datetime
from collections import defaultdict
from dataclasses import dataclass
from functools import lru_cache
from io import BytesIO
from typing import Iterable, Union

from devind_helpers.orm_utils import get_object_or_404
from devind_helpers.schema.types import ErrorFieldType
from devind_helpers.utils import gid2int
from django.db.models import Max, Q
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from apps.core.models import User
from apps.dcis.models import Cell, Document, Period, Sheet, Status, Value
from apps.dcis.permissions import can_add_document


DIVISION_NAME: str = 'idlistedu'  # Идентификатор поля для получения идентификатора дивизиона
ValueType: Union = Union[str, int, float, bool, datetime.datetime]


@dataclass
class CellData:
    """Ячейка с данными из файла."""
    position: str
    value: ValueType
    column_index: int
    row_index: int
    default_value: ValueType
    sheet_id: int
    column_id: int
    row_id: int
    cell_id: id
    editable: bool


class ExcelReaderSheets:
    """Чтение листов."""

    def __init__(self, filename: str | BytesIO) -> None:
        self.work_book: Workbook = load_workbook(filename)

    @property
    @lru_cache
    def sheet_names(self):
        """Свойство получения списка листов."""
        return self.work_book.sheetnames

    @lru_cache
    def get_headers(self, sheet_name: str) -> list[str]:
        """Функция получения заголовка листа sheet_name."""
        sheet: Worksheet = self.work_book[sheet_name]
        return [sheet.cell(1, j + 1).value for j in range(sheet.max_column)]

    def items(self, sheet_name: str) -> Iterable[dict]:
        """Перечисление элементов начиная со второй строки."""
        sheet: Worksheet = self.work_book[sheet_name]
        headers: list[str] = self.get_headers(sheet_name)
        for i in range(2, sheet.max_row + 1):
            yield {
                headers[j]: self.get_value(sheet.cell(i, j + 1).value)
                for j in range(sheet.max_column)
            }

    def column_items(self, sheet_name: str, field: str) -> list:
        sheet: Worksheet = self.work_book[sheet_name]
        headers: list[str] = self.get_headers(sheet_name)
        column_index: int = headers.index(field) + 1
        return [self.get_value(sheet.cell(i, column_index).value) for i in range(2, sheet.max_row + 1)]

    @staticmethod
    def get_value(value):
        """Получение правильного значения ячейки."""
        if type(value) == str:
            return value.strip()
        elif type(value) == datetime.datetime or type(value) == datetime.date:
            return value.strftime('%Y-%m-%d')
        return value


def add_document_data(
    user: User,
    period_id: str | int,
    file: str | BytesIO,
    status_id: str | int
) -> tuple[list[Document] | None, list[ErrorFieldType]]:
    """Функция для создания документов."""
    period: Period = get_object_or_404(Period, pk=gid2int(period_id))
    status: Status = get_object_or_404(Status, pk=status_id)
    can_add_document(user, period, status, None)
    reader: ExcelReaderSheets = ExcelReaderSheets(file)
    # 1. Проверяем пропуски листов
    sheets, mismatch_sheets = get_sheet(period, reader)
    if mismatch_sheets:
        return None, [ErrorFieldType('file', [f'Несовпадение листов: {", ".join(mismatch_sheets)}'])]
    # 2. Проверяем пропуски дивизионов (организаций или департаментов)
    divisions_id: dict[int, bool] = {}
    for sheet_name in reader.sheet_names:
        if DIVISION_NAME not in reader.get_headers(sheet_name):
            return None, [ErrorFieldType('file', [f'В листе {sheet_name} не найдено поле {DIVISION_NAME}'])]
        divisions_id.update({division_id: True for division_id in reader.column_items(sheet_name, DIVISION_NAME)})
    divisions: dict[int, str] = {
        d['pk']: d['name']
        for d in period.project.division.objects.filter(pk__in=divisions_id.keys()).values('pk', 'name')
    }
    for division in divisions:
        divisions_id[division] = False
    mismatch_divisions: list[str] = [str(division_id) for division_id, freq in divisions_id.items() if freq]
    if mismatch_divisions:
        return None, [ErrorFieldType('file', [f'Следующие дивизионы не найдены: {", ".join(mismatch_divisions)}'])]
    exist_divisions: list[int] = period.division_set.values_list('object_id', flat=True)
    for exist_division in exist_divisions:
        divisions_id[exist_division] = True
    for income_division, exist in divisions_id.items():
        if not exist:
            period.division_set.create(object_id=income_division)
    # 3. Строим структуру данных
    documents_data = get_documents_data(reader, sheets, DIVISION_NAME)
    # 4. Создаем документы
    documents: list[Document] = add_documents(
        user,
        period,
        sheets,
        status,
        documents_data,
        divisions
    )
    return documents, []


def get_sheet(period: Period, reader: ExcelReaderSheets) -> tuple[dict[str, Sheet], list[str]]:
    """Получаем список листов из базы данных и листы, которые не совпали."""
    sheet_names: dict[str, bool] = {sheet_name: True for sheet_name in reader.sheet_names}
    sheets: dict[str, Sheet] = {sheet.name: sheet for sheet in period.sheet_set.all()}
    for sheet in sheets:
        sheet_names[sheet] = False
    return sheets, [sheet_name for sheet_name, freq in sheet_names.items() if freq]


def get_documents_data(
    reader: ExcelReaderSheets,
    sheets: dict[str, Sheet],
    division_name: str
) -> dict[int, dict[str, list[CellData]]]:
    """Формирование структуры данных из ридера.

    {
        division_id: {
            sheet_name: [CellData]
        }
    }
    """
    document_data: dict[int, dict[str, list[CellData]]] = defaultdict(dict)
    for sheet_name in reader.sheet_names:  # По листам
        cells_filter: Q = Q()  # Q(column__sheet=sheets[sheet_name])
        for header in reader.get_headers(sheet_name):
            if header == division_name:
                continue
            column_index, row_index = get_coordinate(header)
            cells_filter |= Q(column__index=column_index, row__index=row_index)
        cells: dict[str, dict[str, ...]] = {
            get_position(cell['column__index'], cell['row__index']): cell
            for cell in Cell.objects.filter(Q(column__sheet_id=sheets[sheet_name].id) & cells_filter).values(
                'id', 'column__index', 'row__index', 'editable', 'default', 'column_id', 'row_id', 'column__sheet_id'
            )
        }
        for value in reader.items(sheet_name):
            division_id: int | None = value.pop(division_name)
            document_data[division_id][sheet_name] = [
                CellData(
                    position=cell_position,
                    value=cell_value,
                    column_index=cells[cell_position]['column__index'],
                    row_index=cells[cell_position]['row__index'],
                    default_value=cells[cell_position]['default'],
                    sheet_id=cells[cell_position]['column__sheet_id'],
                    column_id=cells[cell_position]['column_id'],
                    row_id=cells[cell_position]['row_id'],
                    cell_id=cells[cell_position]['id'],
                    editable=cells[cell_position]['editable']
                ) for cell_position, cell_value in value.items()
            ]
    return document_data


def add_documents(
    user: User,
    period: Period,
    sheets: dict[str, Sheet],
    status: Status,
    documents_data: dict[int, dict[str, list[CellData]]],
    divisions: dict[int, str]
) -> list[Document]:
    """Создание и запись документов со значениями."""
    max_versions: dict[int, int] = {
        version['object_id']: version['version__max'] + 1
        for version in Document.objects.filter(
            period=period,
            object_id__in=divisions.keys()
        ).values('object_id').annotate(Max('version')).order_by()
    }
    documents: list[Document] = []
    values: list[Value] = []
    for division_id, sheets_data in documents_data.items():
        if max_versions.get(division_id, 1) > 1 and not period.versioning:
            continue
        document: Document = Document.objects.create(
            version=max_versions.get(division_id, 1),
            updated_by=user,
            user=user,
            period=period,
            object_id=division_id,
            object_name=divisions[division_id]
        )
        document.sheets.add(*sheets.values())
        document.documentstatus_set.create(status=status, user=user)
        cell_data: CellData
        values.extend(
            [
                Value(
                    value=cell_data.value,
                    document=document,
                    sheet=sheets[sheet_name],
                    column_id=cell_data.column_id,
                    row_id=cell_data.row_id
                )
                for sheet_name, cells_data in sheets_data.items()
                for cell_data in cells_data if check_value(cell_data)
            ]
        )
        documents.append(document)
    Value.objects.bulk_create(values)
    return documents


def check_value(cell_data: CellData) -> bool:
    """Проверяем на возможность добавления значения."""
    if cell_data.value == cell_data.default_value:
        return False
    if str(cell_data.value):
        return True
    return False


def get_coordinate(position: str) -> tuple[int, int]:
    """Получение целочисленных индексов позиции."""
    column_letter, row = coordinate_from_string(position)
    return column_index_from_string(column_letter), row


def get_position(column: int, row: int) -> str:
    """Получаем позицию документа.

        - 1, 1 -> A1
    """
    return f'{get_column_letter(column)}{row}'
