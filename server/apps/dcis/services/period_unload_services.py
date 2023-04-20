"""Модуль, отвечающий за выгрузку периода."""

from dataclasses import dataclass
from datetime import datetime
from typing import Callable

from devind_dictionaries.models import Organization
from django.conf import settings
from django.core.exceptions import ValidationError
from django.db.models import QuerySet
from openpyxl.cell.cell import VALID_TYPES
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from apps.core.models import User
from apps.dcis.helpers.exceptions import is_raises
from apps.dcis.models import Cell, Document, MergedCell, Period, Sheet, Value
from apps.dcis.models.sheet import KindCell
from apps.dcis.permissions import can_view_period_result
from apps.dcis.services.divisions_services import get_user_period_divisions


@dataclass
class CellData:
    """Данные ячейки."""
    value: str
    data_type: str = 's'
    number_format: str | None = None


@dataclass
class DataSource:
    """Источник данных для ячейки."""
    organization: Organization
    document: Document | None
    cells: list[CellData]


@dataclass
class UnloadHeaderCell:
    """Ячейка заголовка столбца в выгрузке."""
    name: str
    start_row: int | None = None
    start_column: int | None = None
    end_row: int | None = None
    end_column: int | None = None


@dataclass
class Column:
    """Выгружаемый столбец."""
    get_cell_data: Callable[[DataSource], CellData]
    cells: list[UnloadHeaderCell]


@dataclass
class HeaderCell:
    """Ячейка в шапке таблицы."""
    cell: Cell
    merged_cell: MergedCell | None


@dataclass
class CellGroups:
    """Группы ячеек.
    - value_cells - крайняя правая прямоугольная группа ячеек без readonly и объединенных ячеек
    - column_header_cell - ячейки в шапке таблицы
    - row_header_cells - столбец примыкающий слева к value_cells
    """
    value_cells: list[Cell]
    column_header_cells: list[HeaderCell]
    row_header_cells: list[Cell]


class PeriodUnload:
    """Выгрузка периода в формате Excel."""

    def __init__(
        self,
        user: User,
        period: Period,
        organization_ids: list[int],
        organization_kinds: list[str],
        status_ids: list[int],
        unload_curator_group: bool,
        unload_financing_paragraph: bool,
        unload_without_document: bool,
        unload_default: bool,
        apply_number_format: bool,
        unload_heads: bool,
        unload_children: bool,
        empty_cell: str
    ) -> None:
        """Инициализация.
        - user - текущий пользователь
        - period - выгружаемый период
        - organization_ids - фильтрация по идентификаторам организаций
        - organization_kinds - типы организаций
        - status_ids - фильтрация по идентификаторам статусов
        - unload_curator_group - выгружать кураторскую группу
        - unload_financing_paragraph - выгружать параграф финансирования
        - unload_without_document - выгружать организации без документов
        - unload_default - выгружать значение по умолчанию при отсутствии значения в документе
        - apply_number_format - применять числовой формат
        - unload_heads - выгружать листы для головных учреждений
        - unload_children - выгружать листы для филиалов
        - empty_cell - строка в пустой ячейке
        """
        self.period = period
        self.path = settings.DOCUMENTS_DIR / f'document_{datetime.now().strftime("%d-%m-%Y_%H-%M-%S")}.xlsx'
        self._organizations = self._filter_organizations(user, period, organization_ids, organization_kinds)
        self._status_ids = status_ids
        self._unload_curator_group = unload_curator_group
        self._unload_financing_paragraph = unload_financing_paragraph
        self._unload_without_document = unload_without_document
        self._unload_default = unload_default
        self._apply_number_format = apply_number_format
        if unload_heads and unload_children:
            self._sheets = self.period.sheet_set.all()
        else:
            self._sheets = self.period.sheet_set.filter(show_head=unload_heads, show_child=unload_children)
        self._empty_cell = empty_cell
        self._documents = Document.objects.filter(
            period=self.period,
            object_id__in=self._organizations.values_list('id', flat=True)
        )
        self._documents_map: dict[int, Document | None] | None = None

    @property
    def documents_map(self) -> dict[int, Document | None]:
        """Отображение идентификаторов организаций на документы."""
        if self._documents_map is None:
            self._documents_map = self._build_documents_map()
        return self._documents_map

    def unload(self) -> str:
        """Выгрузка."""
        workbook = Workbook()
        workbook.remove(workbook.active)
        if self._sheets.count() == 0:
            raise ValidationError('Ни один лист периода не соответствует настройкам выгрузки.')
        for sheet in self._sheets:
            worksheet = workbook.create_sheet(sheet.name)
            cells = self._get_cells(sheet)
            cell_groups = self._get_cell_groups(sheet, cells)
            columns = self._build_columns(sheet, cell_groups)
            self._save_columns(worksheet, columns)
            self._save_rows(worksheet, sheet, columns, cell_groups)
        workbook.save(self.path)
        return f'/{self.path.relative_to(settings.BASE_DIR)}'

    @staticmethod
    def _filter_organizations(
        user: User,
        period: Period,
        organization_ids: list[int],
        organization_kinds: list[str],
    ) -> QuerySet[Organization]:
        """Фильтрация организаций."""
        organizations = get_user_period_divisions(user, period)
        if len(organization_kinds):
            organization_kinds_values = []
            for organization_kind in organization_kinds:
                if organization_kind == 'Отсутствует':
                    organization_kinds_values.extend(['', None])
                else:
                    organization_kinds_values.append(organization_kind)
            organizations = organizations.filter(attributes__org_type__in=organization_kinds_values)
        if len(organization_ids):
            organizations = organizations.filter(id__in=organization_ids)
        return organizations

    @classmethod
    def _save_columns(cls, worksheet: Worksheet, columns: list[Column]) -> None:
        """Сохранение название столбцов на лист Excel."""
        max_row_count = cls._get_header_size(columns)
        for i, column in enumerate(columns, 1):
            for column_cell in column.cells:
                if column_cell.start_row is None:
                    column_cell.start_row = 1
                    column_cell.start_column = i
                    column_cell.end_row = max_row_count
                    column_cell.end_column = i
                cell = worksheet.cell(row=column_cell.start_row, column=column_cell.start_column, value=column_cell.name)
                cell.alignment = Alignment(vertical='top')
                if column_cell.start_row != column_cell.end_row or column_cell.start_column != column_cell.end_column:
                    worksheet.merge_cells(
                        start_row=column_cell.start_row,
                        start_column=column_cell.start_column,
                        end_row=column_cell.end_row,
                        end_column=column_cell.end_column,
                    )

    def _save_rows(self, worksheet: Worksheet, sheet: Sheet, columns: list[Column], cell_groups: CellGroups) -> None:
        """Сохранение строки на лист Excel."""
        row_index = self._get_header_size(columns) + 1
        values = Value.objects.filter(sheet=sheet)
        for organization in self._get_organizations(sheet):
            document = self.documents_map[organization.id]
            data_source = DataSource(
                organization=organization,
                document=document,
                cells=self._get_cells_data(cell_groups.value_cells, values, document)
            )
            for column_index, column in enumerate(columns, 1):
                cell_data = column.get_cell_data(data_source)
                cell = worksheet.cell(row=row_index, column=column_index, value=cell_data.value)
                if cell_data.data_type != KindCell.DATE or cell_data.value is not None:
                    cell.data_type = cell_data.data_type
                if cell_data.number_format:
                    cell.number_format = cell_data.number_format
            row_index += 1

    def _build_documents_map(self) -> dict[int, Document | None]:
        """Отображение идентификаторов организаций на документы."""
        result: dict[int, Document | None] = {}
        for organization in self._organizations:
            result[organization.id] = next(
                (document for document in self._documents if organization.id == document.object_id),
                None
            )
        return result

    def _build_columns(self, sheet: Sheet, cell_groups: CellGroups) -> list[Column]:
        """Построение столбцов."""
        left_columns = self._build_left_columns()
        data_columns = self._build_data_columns(sheet, cell_groups, len(left_columns) + 1)
        right_columns = self._build_right_columns()
        return [*left_columns, *data_columns, *right_columns]

    def _build_left_columns(self) -> list[Column]:
        """Получение столбцов слева от данных."""
        columns = [
            Column(
                get_cell_data=lambda s: CellData(s.organization.attributes['idlistedu'], KindCell.NUMERIC),
                cells=[UnloadHeaderCell('IdListEdu')],
            ),
            Column(
                get_cell_data=lambda s: CellData(s.organization.parent.attributes['idlistedu']
                if s.organization.parent else None, KindCell.NUMERIC),
                cells=[UnloadHeaderCell(name='id_parent')],
            ),
            Column(
                get_cell_data=lambda s: CellData(s.organization.kodbuhg, KindCell.NUMERIC),
                cells=[UnloadHeaderCell(name='Бухкод')],
            ),
            Column(
                get_cell_data=lambda s: CellData(
                    s.organization.region.common_id if s.organization.region else None,
                    KindCell.NUMERIC
                ),
                cells=[UnloadHeaderCell(name='Код региона')],
            ),
            Column(
                get_cell_data=lambda s: CellData(s.organization.region.name if s.organization.region else None),
                cells=[UnloadHeaderCell(name='Регион')],
            ),
            Column(
                get_cell_data=lambda s: CellData(s.organization.name),
                cells=[UnloadHeaderCell(name='Название учреждения')],
            ),
            Column(
                get_cell_data=lambda s: CellData(s.organization.parent.name if s.organization.parent else None),
                cells=[UnloadHeaderCell(name='Название головного учреждения')],
            ),
            Column(
                get_cell_data=lambda s: CellData(s.document.last_status.status.name
                if s.document and s.document.last_status else None),
                cells=[UnloadHeaderCell(name='Статус документа')],
            ),
            Column(
                get_cell_data=lambda s: CellData(s.organization.attributes['org_type']),
                cells=[UnloadHeaderCell(name='Тип организации')],
            ),
        ]
        if self._unload_curator_group:
            def get_curator_group(s: DataSource) -> CellData:
                curator_group = s.organization.curatorgroup_set.first()
                return CellData(curator_group.name if curator_group else None)

            columns.append(Column(get_cell_data=get_curator_group, cells=[UnloadHeaderCell(name='Кураторская группа')]))
        if self._unload_financing_paragraph:
            columns.append(Column(
                get_cell_data=lambda s: CellData(
                    f'{s.organization.attributes["id_paragraph"]} {s.organization.attributes["paragraph"]}'
                ),
                cells=[UnloadHeaderCell(name='Параграф финансирования')]
            ))
        return columns

    @classmethod
    def _build_data_columns(cls, sheet: Sheet, cell_groups: CellGroups, left_shift: int) -> list[Column]:
        """Получение столбцов с данными."""
        row_header_chunk = len(cell_groups.row_header_cells)
        if len(cell_groups.value_cells) % row_header_chunk != 0:
            raise ValidationError(
                f'Количество ячеек со значениями на листе "{sheet.name}" не соответствует числу заголовков строк. '
                 'Проверьте правильность расстановки ячеек только для чтения.'
            )
        unload_header_cells = cls._build_unload_header_cells(cell_groups, left_shift)
        columns: list[Column] = []
        for i, _ in enumerate(cell_groups.value_cells):
            cells = [c for c in unload_header_cells if c.start_column == i + left_shift]
            columns.append(Column(get_cell_data=cls._make_get_cell(i), cells=cells))
        return columns

    @classmethod
    def _build_unload_header_cells(
        cls,
        cell_groups: CellGroups,
        left_shift: int
    ) -> list[UnloadHeaderCell]:
        """Построение ячеек заголовков столбцов в выгрузке."""
        row_header_chunk = len(cell_groups.row_header_cells)
        min_row_index = min((hc.cell.row.index for hc in cell_groups.column_header_cells), default=0)
        if min_row_index == 0:
            unload_header_cells: list[UnloadHeaderCell] = []
            for _ in range(len(cell_groups.value_cells) // row_header_chunk):
                for rh in cell_groups.row_header_cells:
                    unload_header_cells.append(cls._cast_row_header_to_unload(rh, left_shift, 1))
                    left_shift += 1
            return unload_header_cells
        min_row_header_cells = sorted(
            (hc for hc in cell_groups.column_header_cells if hc.cell.row.index == min_row_index),
            key=lambda hc: hc.cell.column.index
        )
        return cls._cast_header_cells_to_unload(cell_groups, 1, min_row_header_cells, left_shift, 1)

    @classmethod
    def _find_cell_children(
        cls,
        cell_groups: CellGroups,
        header_cell: HeaderCell,
        left_shift: int,
        top_shift: int,
    ) -> list[UnloadHeaderCell]:
        """Поиск дочерних ячеек для ячейки `header_cell`."""
        vertical_merge_count = cls._get_vertical_merge_count(header_cell)
        top_shift += vertical_merge_count
        min_column_index = header_cell.cell.column.index
        if header_cell.merged_cell is None:
            max_column_index = header_cell.cell.column.index
        else:
            max_column_index = header_cell.merged_cell.max_col
        children_header_cells: list[HeaderCell] = []
        for hc in cell_groups.column_header_cells:
            if (
                min_column_index <= hc.cell.column.index <= max_column_index and
                header_cell.cell.row.index + vertical_merge_count == hc.cell.row.index
            ):
                children_header_cells.append(hc)
        children_header_cells = sorted(children_header_cells, key=lambda h: h.cell.column.index)
        return cls._cast_header_cells_to_unload(
            cell_groups,
            cls._get_horizontal_merge_count(header_cell),
            children_header_cells,
            left_shift,
            top_shift,
        )

    @classmethod
    def _cast_header_cells_to_unload(
        cls,
        cell_groups: CellGroups,
        horizontal_merge_count: int,
        header_cells: list[HeaderCell],
        left_shift: int,
        top_shift: int,
    ) -> list[UnloadHeaderCell]:
        """Приведение ячеек заголовков столбцов и строк к ячейкам заголовков столбцов в выгрузке."""
        row_header_chunk = len(cell_groups.row_header_cells)
        unload_header_cells: list[UnloadHeaderCell] = []
        for hc in header_cells:
            unload_header_cells.append(cls._cast_column_header_to_unload(cell_groups, hc, left_shift, top_shift))
            unload_header_cells.extend(cls._find_cell_children(cell_groups, hc, left_shift, top_shift))
            left_shift += row_header_chunk * cls._get_horizontal_merge_count(hc)
        cells_horizontal_merge_count = sum(map(cls._get_horizontal_merge_count, header_cells))
        for _ in range(horizontal_merge_count - cells_horizontal_merge_count):
            for rh in cell_groups.row_header_cells:
                unload_header_cells.append(cls._cast_row_header_to_unload(rh, left_shift, top_shift))
                left_shift += 1
        return unload_header_cells

    @classmethod
    def _cast_column_header_to_unload(
        cls,
        cell_groups: CellGroups,
        header_cell: HeaderCell,
        left_shift: int,
        top_shift: int,
    ) -> UnloadHeaderCell:
        """Приведение ячейки заголовка столбца к ячейке заголовка столбца в выгрузке."""
        row_header_chunk = len(cell_groups.row_header_cells)
        vertical_merge_count = cls._get_vertical_merge_count(header_cell)
        horizontal_merge_count = cls._get_horizontal_merge_count(header_cell)
        return UnloadHeaderCell(
            name=header_cell.cell.default_error or header_cell.cell.default,
            start_row=top_shift,
            start_column=left_shift,
            end_row=top_shift + vertical_merge_count - 1,
            end_column=left_shift + row_header_chunk * horizontal_merge_count - 1,
        )

    @classmethod
    def _cast_row_header_to_unload(cls, header_cell: Cell, left_shift: int, top_shift: int) -> UnloadHeaderCell:
        """Приведение ячейки заголовка строки к ячейке заголовка столбца в выгрузке."""
        return UnloadHeaderCell(
            name=header_cell.default_error or header_cell.default,
            start_row=top_shift,
            start_column=left_shift,
            end_row=top_shift,
            end_column=left_shift,
        )

    @staticmethod
    def _get_vertical_merge_count(header_cell: HeaderCell) -> int:
        """Получение числа объединенных ячеек по вертикали."""
        if header_cell.merged_cell is None:
            vertical_merge_count = 1
        else:
            vertical_merge_count = header_cell.merged_cell.max_row - header_cell.merged_cell.min_row + 1
        return vertical_merge_count

    @staticmethod
    def _get_horizontal_merge_count(header_cell: HeaderCell) -> int:
        """Получение числа объединенных ячеек по горизонтали."""
        if header_cell.merged_cell is None:
            horizontal_merge_count = 1
        else:
            horizontal_merge_count = header_cell.merged_cell.max_col - header_cell.merged_cell.min_col + 1
        return horizontal_merge_count

    @classmethod
    def _build_right_columns(cls) -> list[Column]:
        """Получение столбцов справа от данных."""
        return [
            Column(
                get_cell_data=lambda s: CellData(
                    s.document.updated_at.replace(tzinfo=None) if s.document else None,
                    KindCell.DATE,
                    'dd/mm/yyyy\ hh:mm',
                ),
                cells=[UnloadHeaderCell(name='Дата последнего редактирования')],
            ),
            Column(
                get_cell_data=lambda s: CellData(
                    cls._format_user(s.document.updated_by)
                    if s.document and s.document.updated_by else None
                ),
                cells=[UnloadHeaderCell(name='Пользователь')],
            ),
        ]

    def _get_organizations(self, sheet: Sheet) -> list[Organization]:
        """Получение организаций с фильтрацией."""
        result: list[Organization] = []
        parent_filter: dict[str, bool] = {}
        if sheet.show_head and not sheet.show_child:
            parent_filter['parent__isnull'] = True
        elif not sheet.show_head and sheet.show_child:
            parent_filter['parent__isnull'] = False
        for organization in self._organizations.filter(**parent_filter):
            document = self.documents_map[organization.id]
            if document is None and not self._unload_without_document:
                break
            if not len(self._status_ids) or (
                document and document.last_status and document.last_status.status.id in self._status_ids
            ):
                result.append(organization)
        return result

    @staticmethod
    def _get_header_size(columns: list[Column]) -> int:
        """Размер шапки таблицы."""
        return max(len(column.cells) for column in columns)

    @staticmethod
    def _get_cells(sheet: Sheet) -> QuerySet[Cell]:
        """Получение ячеек листа."""
        return Cell.objects.filter(
            row__sheet=sheet,
            row__parent__isnull=True
        ).order_by('row__index', 'column__index')

    @staticmethod
    def _get_merged_cells(sheet: Sheet) -> QuerySet[MergedCell]:
        """Получение объединенных ячеек листа."""
        return MergedCell.objects.filter(sheet=sheet)

    def _get_cells_data(self, cells: list[Cell], values: QuerySet[Value], document: Document) -> list[CellData]:
        """Получение данных для ячеек."""
        result: list[CellData] = []
        for cell in sorted(cells, key=lambda c: [c.column.index, c.row.index]):
            val = cell.default_error or cell.default or self._empty_cell if self._unload_default else self._empty_cell
            if val == '':
                val = None
            for value in values:
                if value.document == document and value.column == cell.column and value.row == cell.row:
                    val = value.value
            data_type = cell.kind
            if data_type not in VALID_TYPES:
                data_type = KindCell.STRING
            result.append(CellData(
                value=val,
                data_type=data_type,
                number_format=cell.number_format if self._apply_number_format else None,
            ))
        return result

    @classmethod
    def _get_cell_groups(cls, sheet: Sheet, cells: QuerySet[Cell]) -> CellGroups:
        """Получение групп ячеек."""
        header_cells: list[HeaderCell] = []
        value_cells: list[Cell] = []
        merged_cells = cls._get_merged_cells(sheet)
        merged_cells_columns: set[int] = set()
        merged_cells_rows: set[int] = set()
        for cell in cells:
            position = f'{get_column_letter(cell.column.index)}{cell.row.index}'
            merged_cell = next((mc for mc in merged_cells if position in mc.cells or position == mc.target), None)
            if merged_cell is not None:
                merged_cells_columns.update({*range(merged_cell.min_col, merged_cell.max_col + 1)})
                merged_cells_rows.update({*range(merged_cell.min_row, merged_cell.max_row + 1)})
                if merged_cell.target == position:
                    header_cells.append(HeaderCell(cell=cell, merged_cell=merged_cell))
            elif not cell.editable:
                header_cells.append(HeaderCell(cell=cell, merged_cell=None))
            else:
                value_cells.append(cell)
        value_cells_copy = [*value_cells]
        value_cells = []
        for cell in value_cells_copy:
            if cell.column.index in merged_cells_columns and cell.row.index in merged_cells_rows:
                header_cells.append(HeaderCell(cell=cell, merged_cell=None))
            else:
                value_cells.append(cell)
        if len(value_cells) == 0:
            raise ValidationError(
                f'Не удалось найти ячейки со значениями на листе "{sheet.name}". '
                 'Проверьте правильность расстановки ячеек только для чтения.'
            )
        cell_groups = cls._normalize_cell_groups(value_cells, header_cells)
        if len(cell_groups.row_header_cells) == 0:
            raise ValidationError(
                f'Не удалось найти ячейки с заголовками строк на листе "{sheet.name}". '
                 'Проверьте правильность расстановки ячеек только для чтения.'
            )
        return cell_groups

    @classmethod
    def _normalize_cell_groups(cls, value_cells: list[Cell], header_cells: list[HeaderCell]) -> CellGroups:
        """Нормализация групп ячеек."""
        column_header_index, row_header_index = cls._cut_excess_headers(value_cells, header_cells)
        column_header_cells: list[HeaderCell] = [hc for hc in header_cells if hc.cell.column.index > row_header_index]
        row_header_cells: list[Cell] = [
            hc.cell for hc in header_cells
            if hc.cell.column.index == row_header_index and hc.cell.row.index > column_header_index
        ]
        return CellGroups(
            value_cells=value_cells,
            column_header_cells=column_header_cells,
            row_header_cells=row_header_cells
        )

    @classmethod
    def _cut_excess_headers(cls, value_cells: list[Cell], header_cells: list[HeaderCell]) -> tuple[int, int]:
        """Отрезание лишних ячеек от ячеек с данными."""
        value_cells_copy: list[Cell] = [*value_cells]
        value_cells.clear()
        row, column = max(c.row.index for c in value_cells_copy), max(c.column.index for c in value_cells_copy)
        row_size, column_size = 0, 0
        while row > 0 and column > 0:
            row_cells = [c for c in value_cells_copy if c.row.index == row and c.column.index >= column]
            column_cells = [c for c in value_cells_copy if c.column.index == column and c.row.index >= row]
            if len(row_cells) == row_size + 1 and len(column_cells) == column_size + 1:
                value_cells.extend(row_cells)
                value_cells.extend([c for c in column_cells if c.row.index != row or c.column.index != column])
                row, column = row - 1, column - 1
                row_size, column_size = row_size + 1, column_size + 1
            elif len(row_cells) == row_size + 1 or len(row_cells) == row_size and next(
                (c for c in row_cells if c.column.index == column),
                None
            ) is None:
                value_cells.extend([c for c in row_cells if c.row.index != row or c.column.index != column])
                row -= 1
                column_size += 1
            elif len(column_cells) == column_size + 1 or len(column_cells) == column_size and next(
                (c for c in column_cells if c.row.index == row),
                None
            ) is None:
                value_cells.extend([c for c in column_cells if c.row.index != row or c.column.index != column])
                column -= 1
                row_size += 1
            else:
                break
        for cell in set(value_cells_copy) - set(value_cells):
            header_cells.append(HeaderCell(cell=cell, merged_cell=None))
        return row, column

    @staticmethod
    def _make_get_cell(i: int) -> Callable[[DataSource], CellData]:
        """Создание функции для получения значения."""
        def get_cell(s: DataSource) -> CellData:
            return s.cells[i]
        return get_cell

    @staticmethod
    def _format_user(user: User) -> str:
        """Форматирование пользователя."""
        result = f'{user.last_name} {user.first_name}'
        if user.sir_name:
            return f'{result} {user.sir_name}'
        return result

    @staticmethod
    def _is_numeric(value: str) -> bool:
        """Является ли значение типом `KindCell.NUMERIC`"""
        return not is_raises(ValueError, lambda: int(value) or float(value))


def unload_period(
    user: User,
    period: Period,
    organization_ids: list[int],
    organization_kinds: list[str],
    status_ids: list[int],
    unload_curator_group: bool,
    unload_financing_paragraph: bool,
    unload_without_document: bool,
    unload_default: bool,
    apply_number_format: bool,
    unload_heads: bool,
    unload_children: bool,
    empty_cell: str
) -> str:
    """Выгрузка периода в формате Excel."""
    can_view_period_result(user, period)
    return PeriodUnload(
        user=user,
        period=period,
        organization_ids=organization_ids,
        organization_kinds=organization_kinds,
        status_ids=status_ids,
        unload_curator_group=unload_curator_group,
        unload_financing_paragraph=unload_financing_paragraph,
        unload_without_document=unload_without_document,
        unload_default=unload_default,
        apply_number_format=apply_number_format,
        unload_heads=unload_heads,
        unload_children=unload_children,
        empty_cell=empty_cell,
    ).unload()
