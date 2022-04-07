from typing import List, Dict, Optional, Type
import posixpath
from dataclasses import dataclass
from datetime import datetime
from os.path import join

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.styles.colors import WHITE
from openpyxl.utils import get_column_letter
from django.conf import settings
from django.db.models import Q

from apps.dcis.models import Project, Period, Cell, Document, ColumnDimension, RowDimension, Value


@dataclass
class BuildRow:
    """Дата класс содержащий строку и основную информацию о строке."""
    row: RowDimension
    row_add_date: str
    row_update_date: str
    division_name: str
    division_header: str
    user: str


@dataclass
class BuildCell:
    """Дата класс содержащий собираемую информацию о ячейки."""
    cell: Cell
    value: str
    alignment: Alignment
    font: Font
    border: Border
    pattern_fill: PatternFill


class DocumentUnload:
    """Выгрузка документа в формате эксель."""

    ALLOW_ADDITIONAL: List[str] = ['row_add_date', 'row_update_date', 'division_name', 'division_header', 'user']

    def __init__(self, document: Document, host: str, additional: List[str], divisions_id=None):
        """Инициализация

            document - выгружаемый документ
            host - текущий хост
            additional - дополнительные параметры
            divisions_id - выгружаемые дивизионы в запросе
        """
        if divisions_id is None:
            divisions_id = []
        self.document: Document = document
        self.period: Period = Period.objects \
            .select_related('project') \
            .get(pk=self.document.period_id)
        self.project: Project = self.period.project
        self.sheets = document.sheets\
            .prefetch_related('columndimension_set', 'mergedcell_set')\
            .all()
        self.host: str = host
        self.additional: List[str] = additional
        self.divisions_id: List[int] = divisions_id
        self.path: str = join(settings.DOCUMENTS_DIR, f'document_{datetime.now().strftime("%d-%m-%Y_%H-%M-%S")}.xlsx')

    def xlsx(self):
        workbook: Workbook = Workbook()
        workbook.remove(workbook.active)
        for sheet in self.sheets:
            work_sheet = workbook.create_sheet(sheet.name)
            columns: List[ColumnDimension] = sheet.columndimension_set.all()
            rows: List[RowDimension] = sheet.rowdimension_set.filter(
                Q(parent__isnull=True, document__isnull=True) | Q(
                    document=self.document, parent_id__isnull=False, object_id__in=self.divisions_id
                )
            )
            rows_id: List[int] = [row.id for row in rows]
            cells: List[Cell] = Cell.objects.filter(row_id__in=rows_id).all()
            values: List[Value] = Value.objects.filter(document=self.document, row_id__in=rows_id).all()
            build_rows: List[BuildRow] = self._build_rows(rows)
            build_cells: Dict[str, BuildCell] = self._build_cells(cells, values)

            # Собираем xlsx файл
            row_index: int = 1
            for build_row in build_rows:
                column_index: int = 1
                for column in columns:
                    cell: BuildCell = build_cells.get(f'{column.pk}:{build_row.row.pk}')

                    column_index += 1

                # Дополнительные строки
                for additionalColumn in self.additional:
                    pass

                if build_row.row.height:
                    work_sheet.row_dimensions[row_index].height = build_row.row.height
                row_index += 1

            for column in columns:
                if column.width:
                    work_sheet.column_dimensions[get_column_letter(column.index)].width = column.width // 7

            # При расчете нужно учитывать на сколько добавлено дочерних строк
            # column_offset: int = 0
            # row_offset: int = 0
            # for merge_cell in sheet.mergedcell_set.all():
            #     work_sheet.merge_cells(
            #         start_column=merge_cell.min_col + column_offset,
            #         start_row=merge_cell.min_row + row_offset,
            #         end_column=merge_cell.max_col + column_offset,
            #         end_row=merge_cell.max_row + row_offset
            #     )
        workbook.save(self.path)
        return posixpath.relpath(self.path, settings.BASE_DIR)

    def _build_rows(self, rows: List[RowDimension], parent_id: Optional[Type[int]] = None) -> List[BuildRow]:
        """Функция собирает все строки, включая дочерние в плоский массив."""
        build_rows: List[BuildRow] = []
        current_rows: List[RowDimension] = [row for row in rows if row.parent_id == parent_id]
        for current_row in current_rows:
            date_format = '%H:%M %d.%m.%Y'
            build_row =BuildRow(
                current_row,
                current_row.created_at.strftime(date_format),
                current_row.updated_at.strftime(date_format),
                '',
                '',
                current_row.user.get_full_name if current_row.user is not None else ''
            )
            build_rows = [*build_rows, build_row, *self._build_rows(rows, current_row.parent_id)]
        return build_rows

    def _build_cells(self, cells: List[Cell], values: List[Value]) -> Dict[str, BuildCell]:
        """Собираем ячейки в хеш таблицу для индексации."""
        build_values: Dict[str, Value] = {f'{value.column_id}:{value.row_id}': value for value in values}
        return {
            f'{cell.column_id}:{cell.row_id}': BuildCell(
                cell,
                build_values.get(f'{cell.column_id}:{cell.row_id}', cell.default),
                self._cell_alignment(cell),
                self._cell_font(cell),
                self._cell_border(cell),
                self._cell_pattern_fill(cell)
            )
            for cell in cells
        }

    def xlsx_old(self):
        """Генерация выгрузки."""
        workbook: Workbook = Workbook()
        for sheet in self.sheets:
            ws = workbook.create_sheet(sheet.name)
            columns = sheet.columndimension_set.all()
            rows = sheet.rowdimension_set.all()
            cells = Cell.objects.filter(column__in=columns).prefetch_related('row', 'column').all()
            values = sheet.value_set.filter(document=self.document).all()
            build_values = {}
            for build_value in values:
                key = f'{build_value.column_id}:{build_value.row_id}'
                build_values[key] = build_value.value
            # Вписываем значения в ячейки
            for cell in cells:
                row_position = cell.row.index
                column_position = cell.column.index
                value = build_values.get(f'{cell.column_id}:{cell.row_id}', cell.default or '')
                ws.cell(row_position, column_position, value).alignment = self._cell_alignment(cell)
                # Стили шрифта ячеек
                ws.cell(row_position, column_position).font = self._cell_font(cell)
                if cell.border_style:
                    ws.cell(row_position, column_position).border = self._cell_border(cell)
                # Заливка ячейки
                if cell.background != '#FFFFFF' and cell.background != '#FFFFFFFF' and cell.background != WHITE:
                    ws.cell(row_position, column_position).fill = self._cell_pattern_fill(cell)
            # Ширина и высота для колонок и строк соответственно
            for column in columns:
                if column.width:
                    ws.column_dimensions[get_column_letter(column.index)].width = column.width // 7
            for row in rows:
                if row.height:
                    ws.row_dimensions[row.index].height = row.height

            # Объединение ячеек
            for mc in sheet.mergedcell_set.all():
                ws.merge_cells(
                    start_row=mc.min_row,
                    start_column=mc.min_col,
                    end_row=mc.max_row,
                    end_column=mc.max_col
                )
        workbook.save(self.path)
        return posixpath.relpath(self.path, settings.BASE_DIR)

    @staticmethod
    def _cell_alignment(cell: Cell) -> Alignment:
        return Alignment(
            vertical=cell.vertical_align if cell.vertical_align != 'middle' else 'center',
            horizontal=cell.horizontal_align,
            wrap_text=True
        )

    @staticmethod
    def _cell_font(cell: Cell) -> Font:
        return Font(
            size=cell.size,
            bold=cell.strong,
            italic=cell.italic,
            strike=cell.strike,
            underline=cell.underline,
            color=f'{cell.color[1:]}'
        )

    def _cell_border(self, cell: Cell) -> Border:
        border_styles = {
            position: self._cell_border_side(cell, position)
            for position in ['top', 'bottom', 'left', 'right', 'diagonal']
        }
        return Border(
            diagonalDown=cell.border_style.get('diagonalDown'),
            diagonalUp=cell.border_style.get('diagonalUp'),
            **border_styles
        )

    @staticmethod
    def _cell_pattern_fill(cell: Cell) -> PatternFill:
        return PatternFill(
            fill_type='solid',
            start_color=f'{cell.background[1:]}',
            end_color=f'{cell.background[1:]}'
        )

    @staticmethod
    def _cell_border_side(cell: Cell, position: str) -> Side:
        return Side(
            border_style=cell.border_style.get(position),
            color=f'{cell.border_color[position][1:]}' if cell.border_color[position] else None
        )
