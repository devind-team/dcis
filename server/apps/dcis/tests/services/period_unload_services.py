"""Тестирование модуля, отвечающего за выгрузку периода."""

import datetime
import os
from dataclasses import dataclass
from itertools import product
from pathlib import Path

from devind_dictionaries.models import Organization, Region
from django.conf import settings
from django.contrib.contenttypes.models import ContentType
from django.test import TestCase
from openpyxl.reader.excel import load_workbook

from apps.core.models import User
from apps.dcis.models import Cell, CuratorGroup, Document, Period, Project, RowDimension, Sheet, Status
from apps.dcis.models.sheet import ColumnDimension, KindCell
from apps.dcis.services.period_unload_services import unload_period


@dataclass
class CellData:
    """Данные ячейки"""
    default: str
    kind: str = KindCell.STRING
    number_format: str | None = None
    editable: bool = False


class UnloadPeriodTestCase(TestCase):
    """Тестирование выгрузки периода."""

    RESOURCES_DIR = settings.BASE_DIR / 'apps' / 'dcis' / 'tests' / 'resources' / 'period_unload_services'

    def setUp(self) -> None:
        """Создание данных для тестирования."""
        self.superuser = User.objects.create(
            username='superuser',
            email='superuser@gmail.com',
            first_name='Иван',
            last_name='Иванов',
            sir_name='Иванович',
            is_superuser=True,
        )

        self.organization_content_type = ContentType.objects.get_for_model(Organization)
        self.project = Project.objects.create(content_type=self.organization_content_type)
        self.period = Period.objects.create(project=self.project)

        self.curator_group = CuratorGroup.objects.create()
        self.curator_group.users.add(self.superuser)

        self.statuses = [Status.objects.create(name=f'Статус №{i}') for i in range(1, 7)]

        self.head_sheet = self.period.sheet_set.create(
            name='Головная форма',
            show_head=True,
            show_child=False,
        )
        self.head_columns, self.head_rows = self._create_cells(self.head_sheet)

        self.child_sheet = self.period.sheet_set.create(
            name='Форма филиала',
            show_head=False,
            show_child=True,
        )
        self.child_columns, self.child_rows = self._create_cells(self.child_sheet)

        self.head_organizations: list[Organization] = []
        for i in range(1, 5):
            organization = Organization.objects.create(
                attributes={
                    'idlistedu': str(i),
                    'org_type': f'ВУЗ{i}',
                },
                kodbuhg=i + 10,
                region=Region.objects.create(name=f'Регион №{i}', common_id=i + 20),
                name=f'Корневая организация №{i}',
            )
            self.period.division_set.create(object_id=organization.id)
            self.head_organizations.append(organization)

        self.curator_group.organization.set(self.head_organizations[:2])

        self.child_organizations: list[Organization] = []
        for i, parent in enumerate(self.head_organizations, 5):
            organization = Organization.objects.create(
                attributes={
                    'idlistedu': str(i),
                    'org_type': f'ВУЗ{i}',
                },
                kodbuhg=i + 10,
                parent=parent,
                region=Region.objects.create(name=f'Регион №{i}', common_id=i + 20),
                name=f'Дочерняя организация №{i}',
            )
            self.period.division_set.create(object_id=organization.id)
            self.child_organizations.append(organization)

        for i, data in enumerate(zip(
            [*self.head_organizations[:-1], *self.child_organizations[:-1]],
            [*([self.head_columns] * 3), *([self.child_columns] * 3)],
            [*([self.head_rows] * 3), *([self.child_rows] * 3)],
            self.statuses
        ), 1):
            organization, columns, rows, status = data
            document = Document.objects.create(period=self.period, updated_by=self.superuser, object_id=organization.id)
            document.documentstatus_set.create(status=status, user=self.superuser)
            self._create_values(
                document,
                columns,
                rows,
                [
                    str(i), f'{i}.5', f'Значение {i}', f'Значение {i + 1}',
                    str(i + 1), f'{i + 1}.5', f'Значение {i + 2}',
                ],
            )
            Document.objects.filter(id=document.id).update(
                updated_at=datetime.datetime(2023, 1, 10 + i, tzinfo=datetime.timezone.utc)
            )

        self.actual_path: str | None = None

    def tearDown(self) -> None:
        """Очистка данных после тестирования."""
        if self.actual_path:
            os.remove(self.actual_path)

    def test_without_filter_admin(self) -> None:
        """Тестирование выгрузки без фильтрации, если пользователь видит все организации периода."""
        expected_path = self.RESOURCES_DIR / 'test_without_filter_admin.xlsx'
        self.actual_path = settings.BASE_DIR / unload_period(**self._get_unload_default_settings())[1:]
        self._assert_worksheets_equal(expected_path, self.actual_path)

    def test_without_filter_curator(self) -> None:
        """Тестирование выгрузки без фильтрации, если пользователь видит не все организации периода."""
        self.superuser.is_superuser = False
        self.superuser.save(update_fields=('is_superuser',))
        expected_path = self.RESOURCES_DIR / 'test_without_filter_curator.xlsx'
        self.actual_path = settings.BASE_DIR / unload_period(**self._get_unload_default_settings())[1:]
        self._assert_worksheets_equal(expected_path, self.actual_path)

    def test_organization_filter(self) -> None:
        """Тестирование выгрузки с фильтрацией по организациям."""
        expected_path = self.RESOURCES_DIR / 'test_organization_filter.xlsx'
        self.actual_path = settings.BASE_DIR / unload_period(**{
            **self._get_unload_default_settings(),
            'organization_ids': [self.head_organizations[0].id, self.head_organizations[1].id],
        })[1:]
        self._assert_worksheets_equal(expected_path, self.actual_path)

    def test_organization_kind_filter(self) -> None:
        """Тестирование выгрузки с фильтрацией по типу организации."""
        expected_path = self.RESOURCES_DIR / 'test_organization_kind_filter.xlsx'
        self.actual_path = settings.BASE_DIR / unload_period(**{
            **self._get_unload_default_settings(),
            'organization_kinds': ['ВУЗ1', 'ВУЗ2'],
        })[1:]
        self._assert_worksheets_equal(expected_path, self.actual_path)

    def test_status_filter(self) -> None:
        """Тестирование выгрузки с фильтрацией по статусам."""
        expected_path = self.RESOURCES_DIR / 'test_status_filter.xlsx'
        self.actual_path = settings.BASE_DIR / unload_period(**{
            **self._get_unload_default_settings(),
            'status_ids': [self.statuses[0].id, self.statuses[2].id],
        })[1:]
        self._assert_worksheets_equal(expected_path, self.actual_path)

    def test_unload_without_document(self) -> None:
        """Тестирование выгрузки организаций без документов."""
        expected_path = self.RESOURCES_DIR / 'test_unload_without_document.xlsx'
        self.actual_path = settings.BASE_DIR / unload_period(**{
            **self._get_unload_default_settings(),
            'unload_without_document': True,
        })[1:]
        self._assert_worksheets_equal(expected_path, self.actual_path)

    def test_unload_default(self) -> None:
        """Тестирование выгрузки значений по умолчанию при отсутствии значений в документе."""
        expected_path = self.RESOURCES_DIR / 'test_unload_default.xlsx'
        self.actual_path = settings.BASE_DIR / unload_period(**{
            **self._get_unload_default_settings(),
            'unload_default': True,
        })[1:]
        self._assert_worksheets_equal(expected_path, self.actual_path)

    def test_apply_number_format(self) -> None:
        """Тестирование применения числового формата."""
        expected_path = self.RESOURCES_DIR / 'test_apply_number_format.xlsx'
        self.actual_path = settings.BASE_DIR / unload_period(**{
            **self._get_unload_default_settings(),
            'apply_number_format': True,
        })[1:]
        self._assert_worksheets_equal(expected_path, self.actual_path)

    def test_unload_children(self) -> None:
        """Тестирование выгрузки листов только для филиалов."""
        expected_path = self.RESOURCES_DIR / 'test_unload_children.xlsx'
        self.actual_path = settings.BASE_DIR / unload_period(**{
            **self._get_unload_default_settings(),
            'unload_heads': False,
            'unload_children': True,
        })[1:]
        self._assert_worksheets_equal(expected_path, self.actual_path)

    def test_unload_heads_and_children(self) -> None:
        """Тестирование выгрузки листов для головных учреждений и филиалов."""
        expected_path = self.RESOURCES_DIR / 'test_unload_heads_and_children.xlsx'
        self.actual_path = settings.BASE_DIR / unload_period(**{
            **self._get_unload_default_settings(),
            'unload_heads': True,
            'unload_children': True,
        })[1:]
        self._assert_worksheets_equal(expected_path, self.actual_path)

    def test_empty_cell(self) -> None:
        """Тестирование выгрузки строки в пустой ячейке."""
        expected_path = self.RESOURCES_DIR / 'test_empty_cell.xlsx'
        self.actual_path = settings.BASE_DIR / unload_period(**{
            **self._get_unload_default_settings(),
            'empty_cell': '-',
        })[1:]
        self._assert_worksheets_equal(expected_path, self.actual_path)

    def _assert_worksheets_equal(self, expected_path: Path, actual_path: Path) -> None:
        """Проверка книг на равенство по значениям, типу данных и числовому формату."""
        expected_wb = load_workbook(filename=expected_path)
        actual_wb = load_workbook(filename=actual_path)
        self.assertEqual(expected_wb.sheetnames, actual_wb.sheetnames)
        for expected_ws, actual_ws in zip(expected_wb.worksheets, actual_wb.worksheets):
            self.assertEqual(expected_ws.title, actual_ws.title)
            self.assertEqual(expected_ws.min_row, actual_ws.min_row)
            self.assertEqual(expected_ws.max_row, actual_ws.max_row)
            self.assertEqual(expected_ws.min_column, actual_ws.min_column)
            self.assertEqual(expected_ws.max_column, actual_ws.max_column)
            self.assertEqual(set(map(tuple, expected_ws.merged_cells)), set(map(tuple, actual_ws.merged_cells)))
            for r1, r2 in zip(expected_ws, actual_ws):
                for c1, c2 in zip(r1, r2):
                    self.assertEqual(c1.data_type, c2.data_type)
                    self.assertEqual(c1.value, c2.value)
                    self.assertEqual(c1.number_format, c2.number_format)
                    self.assertEqual(c1.alignment.vertical, c2.alignment.vertical)

    def _get_unload_default_settings(self) -> dict:
        """Получение настроек выгрузки по умолчанию."""
        return {
            'user': self.superuser,
            'period': self.period,
            'organization_ids': [],
            'status_ids': [],
            'organization_kinds': [],
            'unload_without_document': False,
            'unload_default': False,
            'apply_number_format': False,
            'unload_heads': True,
            'unload_children': False,
            'empty_cell': '',
        }

    @staticmethod
    def _create_cells(sheet: Sheet) -> tuple[list[ColumnDimension], list[RowDimension]]:
        """Создание ячеек для листа."""
        columns: list[ColumnDimension] = []
        rows: list[RowDimension] = []
        for i in range(1, 7):
            columns.append(sheet.columndimension_set.create(index=i))
            rows.append(sheet.rowdimension_set.create(index=i))
        cells_data = [
            [CellData(f'Шапка: {sheet.name}'), CellData(''), CellData(''), CellData(''), CellData(''), CellData('')],
            [
                CellData(f'Номера: {sheet.name}'), CellData(''), CellData(f'Числа: {sheet.name}'),
                CellData(''), CellData(f'Строки: {sheet.name}'), CellData(''),
            ],
            [
                CellData(f'Номер строки: {sheet.name}'), CellData(f'Номер: {sheet.name}'),
                CellData(f'Целые: {sheet.name}'), CellData(f'2 знака: {sheet.name}'),
                CellData(''), CellData(''),
            ],
            [
                CellData('1'), CellData('№1'),
                CellData('100', KindCell.NUMERIC, None, True),
                CellData('111.5', KindCell.NUMERIC, '0.00', True),
                CellData('Значение 100', KindCell.STRING, None, True),
                CellData('Значение 200', KindCell.STRING, None, True),
            ],
            [
                CellData('2'), CellData('№2'),
                CellData('200', KindCell.NUMERIC, None, True),
                CellData('222.25', KindCell.NUMERIC, '0.00', True),
                CellData('Значение 300', KindCell.STRING, None, True),
                CellData('Значение 400', KindCell.STRING, None, True),
            ],
            [CellData(f'Подвал: {sheet.name}'), CellData(''), CellData(''), CellData(''), CellData(''), CellData('')]
        ]
        cells: list[list[Cell]] = []
        for row, row_cell_data in zip(rows, cells_data):
            cells.append([])
            for column, cell_data in zip(columns, row_cell_data):
                cells[-1].append(Cell.objects.create(
                    default=cell_data.default,
                    kind=cell_data.kind,
                    number_format=cell_data.number_format,
                    editable=cell_data.editable,
                    column=column,
                    row=row,
                ))
        sheet.mergedcell_set.create(min_col=1, min_row=1, max_col=6, max_row=1)
        sheet.mergedcell_set.create(min_col=1, min_row=2, max_col=2, max_row=2)
        sheet.mergedcell_set.create(min_col=3, min_row=2, max_col=4, max_row=2)
        sheet.mergedcell_set.create(min_col=5, min_row=2, max_col=6, max_row=3)
        sheet.mergedcell_set.create(min_col=1, min_row=6, max_col=6, max_row=6)
        return columns, rows

    @staticmethod
    def _create_values(
        document: Document,
        columns: list[ColumnDimension],
        rows: list[RowDimension],
        values: list[str]
    ) -> None:
        """Создание значений для документа."""
        for indices, value in zip(product(range(3, 5), range(2, 6)), values):
            i, j = indices
            if i != 4 or j != 5:
                document.value_set.create(
                    value=value,
                    row=rows[i],
                    column=columns[j],
                    sheet=columns[0].sheet,
                )
