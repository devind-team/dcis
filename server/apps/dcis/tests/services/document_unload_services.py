"""Тестирование модуля, отвечающего за выгрузку документа."""

import os
from datetime import datetime
from itertools import product
from pathlib import Path

from devind_dictionaries.models import Department, Organization
from django.conf import settings
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import PermissionDenied
from django.test import TestCase
from openpyxl.reader.excel import load_workbook

from apps.core.models import User
from apps.dcis.models import Cell, ColumnDimension, Document, MergedCell, Period, Project, RowDimension, Value
from apps.dcis.models.sheet import KindCell, Style
from apps.dcis.services.document_unload_services import unload_document


class UnloadDocumentTestCase(TestCase):
    """Тестирование модуля, отвечающего за выгрузку документа."""

    RESOURCES_DIR = settings.BASE_DIR / 'apps' / 'dcis' / 'tests' / 'resources' / 'document_unload_services'

    def setUp(self) -> None:
        """Создание данных для тестирования."""
        self.superuser = User.objects.create(
            username='superuser',
            email='superuser@gmail.com',
            first_name='Иван',
            last_name='Иванов',
            sir_name='Иванович',
            is_superuser=True,
        )
        self.department_head = User.objects.create(
            username='department_head',
            email='department_head@gmail.com',
            first_name='Петр',
            last_name='Петров',
            sir_name='Петрович',
        )
        self.department_employee = User.objects.create(
            username='department_employee',
            email='department_employee@gmail.com',
            first_name='Илья',
            last_name='Ильин',
            sir_name='Ильич',
        )
        self.extra_user = User.objects.create(
            username='extra_user',
            email='extra_user@gmail.com',
            first_name='Василий',
            last_name='Васильев',
            sir_name='Васильевич',
        )

        self.department_content_type = ContentType.objects.get_for_model(Department)
        self.organization_content_type = ContentType.objects.get_for_model(Organization)

        self.department = Department.objects.create(name='Департамент', code=10, user=self.department_head)
        self.department_employee.departments.add(self.department)

        self.head_organization = Organization.objects.create(name='Головная организация', attributes='')
        self.branch_organization = Organization.objects.create(
            name='Филиал',
            parent=self.head_organization,
            attributes='',
        )

        self.border_style = {
            'top': 'thin',
            'left': 'thin',
            'right': 'thin',
            'bottom': 'thin',
            'diagonal': None,
        }
        self.border_color = {
            'top': '#000000',
            'left': '#000000',
            'right': '#000000',
            'bottom': '#000000',
            'diagonal': None,
        }

        self.department_project = Project.objects.create(content_type=self.department_content_type)
        self.department_period = Period.objects.create(project=self.department_project)
        self.department_period.division_set.create(object_id=self.department.id)

        self.department_form1 = self.department_period.sheet_set.create(name='Форма 1', position=0)
        self.department_form1_column_dimensions = [ColumnDimension.objects.create(
            sheet=self.department_form1,
            index=i,
            width=50 * i,
        ) for i in range(1, 4)]
        self.department_form1_row_dimensions = [RowDimension.objects.create(
            sheet=self.department_form1,
            index=i,
            height=25 * i,
            user=self.department_employee,
        ) for i in range(1, 5)]
        for i, dimensions in enumerate(product(
            self.department_form1_column_dimensions,
            self.department_form1_row_dimensions,
        ), 1):
            column, row = dimensions
            Cell.objects.create(
                column=column,
                row=row,
                kind=KindCell.NUMERIC,
                number_format='0.00',
                default=str(i),
                border_style=self.border_style,
                border_color=self.border_color,
                vertical_align=Style.MIDDLE,
                horizontal_align=Style.CENTER,
            )
        MergedCell.objects.create(sheet=self.department_form1, min_col=2, min_row=1, max_col=3, max_row=1)
        MergedCell.objects.create(sheet=self.department_form1, min_col=1, min_row=1, max_col=1, max_row=2)

        self.department_form2 = self.department_period.sheet_set.create(name='Форма 2', position=1)
        self.department_form2_column_dimensions = [ColumnDimension.objects.create(
            sheet=self.department_form2,
            index=i,
            width=100,
        ) for i in range(1, 3)]
        self.department_form2_row_dimensions = [RowDimension.objects.create(
            sheet=self.department_form2,
            index=i,
            height=50,
            user=self.department_employee,
        ) for i in range(1, 5)]
        for i, dimensions in enumerate(product(
            self.department_form2_column_dimensions,
            self.department_form2_row_dimensions,
        )):
            column, row = dimensions
            cell = Cell.objects.create(
                column=column,
                row=row,
                kind=KindCell.TEXT,
                border_style=self.border_style,
                border_color=self.border_color,
            )
            match i:
                case 0:
                    cell.size = 14
                case 1:
                    cell.strong = True
                case 2:
                    cell.italic = True
                case 3:
                    cell.strike = True
                case 4:
                    cell.underline = Style.SINGLE
                case 5:
                    cell.color = '#FF0000'
                case _:
                    cell.background = '#00FF00'
            cell.save()
        for row in [*self.department_form1_row_dimensions, *self.department_form2_row_dimensions]:
            row.created_at = datetime(2022, 9, 21, 13, 30, 0)
            row.save(update_fields=('created_at',))
            RowDimension.objects.filter(id=row.id).update(updated_at=datetime(2022, 9, 21, 13, 30, 0))

        self.departament_document = Document.objects.create(
            period=self.department_period,
            object_id=self.department.id,
            object_name=self.department.name,
        )
        self.departament_document.sheets.set([self.department_form1, self.department_form2])
        Value.objects.create(
            sheet=self.department_form1,
            document=self.departament_document,
            column=self.department_form1_column_dimensions[0],
            row=self.department_form1_row_dimensions[0],
            value='1000',
        )
        for i, dimensions in enumerate(product(
            self.department_form2_column_dimensions,
            self.department_form2_row_dimensions,
        )):
            column, row = dimensions
            Value.objects.create(
                sheet=self.department_form2,
                document=self.departament_document,
                column=column,
                row=row,
                value=f'Значение №{i}'
            )

        self.organization_project = Project.objects.create(content_type=self.organization_content_type)
        self.organization_period = Period.objects.create(project=self.organization_project)
        self.organization_period.division_set.create(object_id=self.head_organization.id)
        self.organization_period.division_set.create(object_id=self.branch_organization.id)

        self.head_organization_form = self.organization_period.sheet_set.create(
            name='Головная форма',
            position=0,
            show_head=True,
            show_child=False,
        )
        self.branch_organization_form = self.organization_period.sheet_set.create(
            name='Форма филиала',
            position=1,
            show_head=False,
            show_child=True,
        )

        for form in [self.head_organization_form, self.branch_organization_form]:
            organization_form_column_dimensions = [ColumnDimension.objects.create(
                sheet=form,
                index=i,
                width=150,
            ) for i in range(1, 5)]
            organization_form_row_dimensions = [RowDimension.objects.create(
                sheet=form,
                index=i,
                height=50,
                dynamic=i == 1,
            ) for i in range(1, 5)]
            for i, dimensions in enumerate(product(
                organization_form_column_dimensions,
                organization_form_row_dimensions,
            )):
                column, row = dimensions
                Cell.objects.create(
                    column=column,
                    row=row,
                    kind=KindCell.TEXT,
                    border_style=self.border_style,
                    border_color=self.border_color,
                    default=f'Значение по умолчанию №{i}'
                )

        self.head_organization_document = Document.objects.create(
            period=self.organization_period,
            object_id=self.head_organization.id,
            object_name=self.head_organization.name,
        )
        self.head_organization_document.sheets.set([self.head_organization_form, self.branch_organization_form])
        Document.objects.filter(id=self.head_organization_document.id).update(
            updated_at=datetime(2022, 9, 21, 13, 30, 0),
        )

        self.branch_organization_document = Document.objects.create(
            period=self.organization_period,
            object_id=self.branch_organization.id,
            object_name=self.branch_organization.name
        )
        self.branch_organization_document.sheets.set([self.head_organization_form, self.branch_organization_form])
        Document.objects.filter(id=self.branch_organization_document.id).update(
            updated_at=datetime(2022, 9, 21, 13, 30, 0),
        )

        Value.objects.create(
            sheet=self.head_organization_form,
            document=self.head_organization_document,
            column=self.head_organization_form.columndimension_set.first(),
            row=self.head_organization_form.rowdimension_set.first(),
            value='Головное значение'
        )
        Value.objects.create(
            sheet=self.branch_organization_form,
            document=self.branch_organization_document,
            column=self.branch_organization_form.columndimension_set.first(),
            row=self.branch_organization_form.rowdimension_set.first(),
            value='Значение филиала',
        )

        self.actual_path: str | None = None

    def tearDown(self) -> None:
        """Очистка данных после тестирования."""
        if self.actual_path:
            os.remove(self.actual_path)

    def test_without_permissions(self) -> None:
        """Тестирование выгрузки документа без разрешений."""
        with self.assertRaises(PermissionDenied):
            unload_document(
                user=self.extra_user,
                document_id=self.departament_document.id,
                additional=[],
            )

    def test_department_without_additional(self) -> None:
        """Тестирование выгрузки документа департамента без дополнительных столбцов."""
        expected_path = self.RESOURCES_DIR / 'test_department_without_additional.xlsx'
        self.actual_path = settings.BASE_DIR / unload_document(
            user=self.superuser,
            document_id=self.departament_document.id,
            additional=[],
        )[1:]
        self._assert_workbooks_equal(expected_path, self.actual_path)

    def test_department_with_wrong_additional(self) -> None:
        """Тестирование выгрузки документа департамента с неправильным дополнительным столбцом."""
        expected_path = self.RESOURCES_DIR / 'test_department_with_wrong_additional.xlsx'
        self.actual_path = settings.BASE_DIR / unload_document(
            user=self.superuser,
            document_id=self.departament_document.id,
            additional=['wrong_additional'],
        )[1:]
        self._assert_workbooks_equal(expected_path, self.actual_path)

    def test_department_with_row_add_date_additional(self) -> None:
        """Тестирование выгрузки документа департамента с датой добавления строки в качестве дополнительного столбца."""
        expected_path = self.RESOURCES_DIR / 'test_department_with_row_add_date_additional.xlsx'
        self.actual_path = settings.BASE_DIR / unload_document(
            user=self.superuser,
            document_id=self.departament_document.id,
            additional=['row_add_date'],
        )[1:]
        self._assert_workbooks_equal(expected_path, self.actual_path)

    def test_department_with_all_additional(self) -> None:
        """Тестирование выгрузки документа департамента со всеми дополнительными столбцами."""
        expected_path = self.RESOURCES_DIR / 'test_department_with_all_additional.xlsx'
        self.actual_path = settings.BASE_DIR / unload_document(
            user=self.superuser,
            document_id=self.departament_document.id,
            additional=['row_add_date', 'row_update_date', 'division_name', 'division_head', 'user'],
        )[1:]
        self._assert_workbooks_equal(expected_path, self.actual_path)

    def test_organization_without_children_rows(self) -> None:
        """Тестирование выгрузки документа организации без подстрок."""
        expected_path = self.RESOURCES_DIR / 'test_organization_without_children_rows.xlsx'
        self.actual_path = settings.BASE_DIR / unload_document(
            user=self.superuser,
            document_id=self.head_organization_document.id,
            additional=[],
        )[1:]
        self._assert_workbooks_equal(expected_path, self.actual_path)

    def test_organization_with_children_rows(self) -> None:
        """Тестирование выгрузки документа организации с подстроками."""
        children_rows = [RowDimension.objects.create(
            sheet=self.head_organization_form,
            index=i,
            height=50,
            parent=self.head_organization_form.rowdimension_set.first(),
            document=self.head_organization_document,
        ) for i in range(1, 3)]
        for i, dimensions in enumerate(product(
            self.head_organization_form.columndimension_set.all(),
            children_rows,
        )):
            column, row = dimensions
            Cell.objects.create(
                column=column,
                row=row,
                kind=KindCell.TEXT,
                border_style=self.border_style,
                border_color=self.border_color,
                default=f'Значение по умолчанию дочерней строки №{i}'
            )
        Value.objects.create(
            sheet=self.head_organization_form,
            document=self.head_organization_document,
            column=self.head_organization_form.columndimension_set.first(),
            row=children_rows[0],
            value='Значение дочерней строки',
        )

        expected_path = self.RESOURCES_DIR / 'test_organization_with_children_rows.xlsx'
        self.actual_path = settings.BASE_DIR / unload_document(
            user=self.superuser,
            document_id=self.head_organization_document.id,
            additional=[],
        )[1:]
        self._assert_workbooks_equal(expected_path, self.actual_path)

    def test_organization_branch(self) -> None:
        """Тестирование выгрузки документа филиала организации."""
        expected_path = self.RESOURCES_DIR / 'test_organization_branch.xlsx'
        self.actual_path = settings.BASE_DIR / unload_document(
            user=self.superuser,
            document_id=self.branch_organization_document.id,
            additional=[],
        )[1:]
        self._assert_workbooks_equal(expected_path, self.actual_path)

    def _assert_workbooks_equal(self, expected_path: Path, actual_path: Path) -> None:
        """Проверка книг Excel на равенство."""
        expected_wb = load_workbook(filename=expected_path)
        actual_wb = load_workbook(filename=actual_path)
        self.assertEqual(expected_wb.sheetnames, actual_wb.sheetnames)
        for expected_ws, actual_ws in zip(expected_wb.worksheets, actual_wb.worksheets):
            self.assertEqual(expected_ws.title, actual_ws.title)
            self.assertEqual(expected_ws.min_row, actual_ws.min_row)
            self.assertEqual(expected_ws.max_row, actual_ws.max_row)
            self.assertEqual(expected_ws.min_column, actual_ws.min_column)
            self.assertEqual(expected_ws.max_column, actual_ws.max_column)
            self.assertEqual(set(map(tuple, expected_ws.merged_cells)), set(map(tuple, actual_ws.merged_cells)))
            for c1, c2 in zip(expected_ws.column_dimensions.values(), actual_ws.column_dimensions.values()):
                self.assertEqual(c1.width, c2.width)
            for r1, r2 in zip(expected_ws.row_dimensions.values(), actual_ws.row_dimensions.values()):
                self.assertEqual(r1.height, r2.height)
            for r1, r2 in zip(expected_ws, actual_ws):
                for c1, c2 in zip(r1, r2):
                    self.assertEqual(c1.data_type, c2.data_type)
                    self.assertEqual(c1.value, c2.value)
                    self.assertEqual(c1.number_format, c2.number_format)
                    self.assertEqual(c1.alignment.vertical, c2.alignment.vertical)
                    self.assertEqual(c1.alignment.horizontal, c2.alignment.horizontal)
                    self.assertEqual(c1.font.size, c2.font.size)
                    self.assertEqual(c1.font.bold, c2.font.bold)
                    self.assertEqual(c1.font.italic, c2.font.italic)
                    self.assertEqual(c1.font.strike, c2.font.strike)
                    self.assertEqual(c1.font.underline, c2.font.underline)
                    self.assertEqual(c1.font.color, c2.font.color)
                    self.assertEqual(c1.border.left, c2.border.left)
                    self.assertEqual(c1.border.top, c2.border.top)
                    self.assertEqual(c1.border.right, c2.border.right)
                    self.assertEqual(c1.border.bottom, c2.border.bottom)
