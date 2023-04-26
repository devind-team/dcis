"""Тесты модуля, отвечающего за работу со значениями."""

from dataclasses import dataclass
from typing import Iterable

from devind_dictionaries.models import Organization
from django.contrib.contenttypes.models import ContentType
from django.db.models import Q
from django.test import TestCase, override_settings
from openpyxl.utils import get_column_letter

from apps.core.models import User
from apps.dcis.helpers.sheet_formula_cache import SheetFormulaContainerCache
from apps.dcis.models import Cell, ColumnDimension, Document, Period, Project, RowDimension, Sheet
from apps.dcis.models.sheet import KindCell, Value
from apps.dcis.services.value_services import ValueInput, recalculate_all_cells, update_or_create_values


@dataclass
class CellData:
    """Данные ячейки"""
    has_value: bool
    value: str
    extra_value: str | None = None
    error: str | None = None


@override_settings(CACHES={'default': {'BACKEND': 'django.core.cache.backends.locmem.LocMemCache'}})
class UpdateOrCreateValuesTestCase(TestCase):
    """Тестирование сервисов для создания или обновления значений."""

    def setUp(self) -> None:
        """Создание данных для тестирования."""
        self.user = User.objects.create(username='user', email='user@gmain.com')

        self.organization_content_type = ContentType.objects.get_for_model(Organization)
        self.project = Project.objects.create(content_type=self.organization_content_type)
        self.period = Period.objects.create(project=self.project)
        self.parent_organization = Organization.objects.create(attributes='')
        self.children_organizations = [
            Organization.objects.create(attributes='', parent=self.parent_organization) for _ in range(3)
        ]
        self.organizations = [self.parent_organization, *self.children_organizations]
        for organization in self.organizations:
            self.period.division_set.create(object_id=organization.id)
        self.parent_document = Document.objects.create(period=self.period, object_id=self.parent_organization.id)
        self.children_documents = [Document.objects.create(
            period=self.period,
            object_id=organization.id
        ) for organization in self.children_organizations]
        self.documents = [self.parent_document, *self.children_documents]

        self.expected_values: dict[Cell, CellData] = {}
        self.forms: list[Sheet] = []
        for f in range(1, 3):
            form = Sheet.objects.create(name=f'Форма{f}', period=self.period)
            columns = [ColumnDimension.objects.create(index=i, sheet=form) for i in range(1, 5)]
            rows = [RowDimension.objects.create(index=i, sheet=form) for i in range(1, 4)]
            cache_container = SheetFormulaContainerCache(name=form.name)
            for i in range(3):
                row = rows[i]
                for j in range(4):
                    column = columns[j]
                    if j == 0:
                        cell = Cell.objects.create(kind=KindCell.NUMERIC, default='1.0', column=column, row=row)
                        for document in self.documents:
                            Value.objects.create(
                                value='2.0',
                                document=document,
                                sheet=form,
                                column=column,
                                row=row,
                            )
                        self.expected_values[cell] = CellData(True, '2.0')
                    elif i != 2 and j == 3:
                        if f == 1:
                            formula = f'=SUM(B{i + 1}:C{i + 1})'
                            cell = Cell.objects.create(kind=KindCell.NUMERIC, formula=formula, column=column, row=row)
                            cache_container.add_formula(f'{get_column_letter(column.index)}{row.index}', formula)
                            for document in self.documents:
                                Value.objects.create(
                                    value='2.0',
                                    document=document,
                                    sheet=form,
                                    column=column,
                                    row=row,
                                )
                            self.expected_values[cell] = CellData(True, '2.0')
                        else:
                            formula = f'=SUM(Форма1!B{i + 1}:D{i + 1})'
                            cell = Cell.objects.create(kind=KindCell.NUMERIC, formula=formula, column=column, row=row)
                            cache_container.add_formula(f'{get_column_letter(column.index)}{row.index}', formula)
                            for document in self.documents:
                                Value.objects.create(
                                    value='4.0',
                                    document=document,
                                    sheet=form,
                                    column=column,
                                    row=row,
                                )
                            self.expected_values[cell] = CellData(True, '4.0')
                    else:
                        cell = Cell.objects.create(kind=KindCell.NUMERIC, default='1.0', column=column, row=row)
                        self.expected_values[cell] = CellData(False, '1.0')
            cache_container.save(sheet_id=form.id)
            self.forms.append(form)

        self.aggregation_form = Sheet.objects.create(name=f'Форма агрегации', period=self.period)
        self.aggregation_row = RowDimension.objects.create(index=1, sheet=self.aggregation_form)
        self.aggregation_columns = [ColumnDimension.objects.create(
            index=i,
            sheet=self.aggregation_form,
        ) for i in range(1, 5)]
        self.aggregation_simple_cell = Cell.objects.create(
            kind=KindCell.NUMERIC,
            default='4.0',
            column=self.aggregation_columns[0],
            row=self.aggregation_row,
        )
        self.aggregation_cell = Cell.objects.create(
            kind=KindCell.NUMERIC,
            default='9.0',
            aggregation=Cell.AGGREGATION_SUM,
            column=self.aggregation_columns[1],
            row=self.aggregation_row,
        )
        self.aggregation_cell.to_cells.create(
            from_cell=Cell.objects.get(column__sheet_id=self.forms[0].id, column__index=1, row__index=3),
        )
        self.aggregation_cell.to_cells.create(
            from_cell=Cell.objects.get(column__sheet_id=self.forms[0].id, column__index=2, row__index=3),
        )
        self.aggregation_aggregation_depends_cell = Cell.objects.create(
            kind=KindCell.NUMERIC,
            default='18.0',
            aggregation=Cell.AGGREGATION_SUM,
            column=self.aggregation_columns[2],
            row=self.aggregation_row,
        )
        self.aggregation_aggregation_depends_cell.to_cells.create(
            from_cell=Cell.objects.get(column__sheet_id=self.forms[0].id, column__index=4, row__index=1),
        )
        self.aggregation_aggregation_depends_cell.to_cells.create(
            from_cell=Cell.objects.get(column__sheet_id=self.forms[1].id, column__index=4, row__index=1),
        )
        self.aggregation_formula_depends_cell = Cell.objects.create(
            kind=KindCell.NUMERIC,
            formula='=0.0',
            default='0.0',
            column=self.aggregation_columns[3],
            row=self.aggregation_row,
        )
        self.aggregation_cache_container = SheetFormulaContainerCache(name=self.aggregation_form.name)
        self.aggregation_cache_container.add_formula(
            f'{get_column_letter(self.aggregation_columns[3].index)}{self.aggregation_row.index}',
            self.aggregation_formula_depends_cell.formula
        )
        self.aggregation_cache_container.save(sheet_id=self.aggregation_form.id)

        self.parent_document.sheets.set([*self.forms, self.aggregation_form])
        for child_document in self.children_documents:
            child_document.sheets.set(self.forms)

    def test_simple_single_create(self) -> None:
        """Тестирование изменения одной ячейки без зависимых формул с созданием значения."""
        form = self.forms[0]
        q = Q(column__sheet_id=form.id, column__index=2, row__index=3)
        cell = Cell.objects.get(q)
        result = update_or_create_values(
            user=self.user,
            document=self.parent_document,
            sheet_id=form.id,
            value_inputs=[ValueInput(cell=cell, value='3.0')],
        )
        self.assertEqual([Value.objects.get(q)], result.values)
        self.assertEqual(CellData(False, '1.0'), self.expected_values[cell])
        self.expected_values[cell] = CellData(True, '3.0')
        self._test_values(self.expected_values, self.parent_document)

    def test_simple_single_update(self) -> None:
        """Тестирование изменения одной без зависимых формул с обновлением значения."""
        form = self.forms[0]
        q = Q(column__sheet_id=form.id, column__index=1, row__index=3)
        cell = Cell.objects.get(q)
        result = update_or_create_values(
            user=self.user,
            document=self.parent_document,
            sheet_id=form.id,
            value_inputs=[ValueInput(cell=cell, value='3.0')],
        )
        self.assertEqual([Value.objects.get(q & Q(document=self.parent_document))], result.values)
        self.assertEqual(CellData(True, '2.0'), self.expected_values[cell])
        self.expected_values[cell] = CellData(True, '3.0')
        self._test_values(self.expected_values, self.parent_document)

    def test_simple_multiple(self) -> None:
        """Тестирование изменения нескольких ячеек без зависимых формул."""
        form = self.forms[0]
        q = Q(column__sheet_id=form.id, column__index__lte=4, row__index=3)
        cells = Cell.objects.filter(q)
        result = update_or_create_values(
            user=self.user,
            document=self.parent_document,
            sheet_id=form.id,
            value_inputs=[ValueInput(cell=cell, value=f'{10 + i:.1f}') for i, cell in enumerate(cells)]
        )
        self.assertEqual(
            set(Value.objects.filter(q & Q(document=self.parent_document))),
            set(result.values)
        )
        for i, cell in enumerate(cells):
            if i == 0:
                self.assertEqual(CellData(True, '2.0'), self.expected_values[cell])
            else:
                self.assertEqual(CellData(False, '1.0'), self.expected_values[cell])
            self.expected_values[cell] = CellData(True, f'{10 + i:.1f}')
        self._test_values(self.expected_values, self.parent_document)

    def test_formula_single(self) -> None:
        """Тестирование изменения одной ячейки с зависимыми формулами."""
        form = self.forms[0]
        q = Q(column__sheet_id=form.id, column__index=2, row__index=1)
        cell = Cell.objects.get(q)
        formula_cell_q = Q(column__sheet_id=self.forms[0].id, column__index=4, row__index=1)
        formula_cell1 = Cell.objects.get(formula_cell_q)
        q |= formula_cell_q
        formula_cell_q = Q(column__sheet_id=self.forms[1].id, column__index=4, row__index=1)
        formula_cell2 = Cell.objects.get(formula_cell_q)
        q |= formula_cell_q
        result = update_or_create_values(
            user=self.user,
            document=self.parent_document,
            sheet_id=form.id,
            value_inputs=[ValueInput(cell=cell, value='3.0')]
        )
        self.assertEqual(CellData(False, '1.0'), self.expected_values[cell])
        self.expected_values[cell] = CellData(True, '3.0')
        self.assertEqual(CellData(True, '2.0'), self.expected_values[formula_cell1])
        self.expected_values[formula_cell1] = CellData(True, '4.0')
        self.assertEqual(CellData(True, '4.0'), self.expected_values[formula_cell2])
        self.expected_values[formula_cell2] = CellData(True, '8.0')
        self.assertEqual(
            set(Value.objects.filter(q & Q(document=self.parent_document))),
            set(result.values)
        )
        self._test_values(self.expected_values, self.parent_document)

    def test_formula_single_error(self) -> None:
        """Тестирование изменения одной ячейки с зависимыми формулами с ошибкой в результате расчета."""
        form = self.forms[0]
        q = Q(column__sheet_id=form.id, column__index=2, row__index=2)
        cell = Cell.objects.get(q)
        formula_cell_q = Q(column__sheet_id=self.forms[0].id, column__index=4, row__index=2)
        formula_cell1 = Cell.objects.get(formula_cell_q)
        q |= formula_cell_q
        formula_cell_q = Q(column__sheet_id=self.forms[1].id, column__index=4, row__index=2)
        formula_cell2 = Cell.objects.get(formula_cell_q)
        q |= formula_cell_q
        formula_cell1.formula = '=SUM(B3:C3) / 0'
        formula_cell1.save(update_fields=('formula',))
        result = update_or_create_values(
            user=self.user,
            document=self.parent_document,
            sheet_id=form.id,
            value_inputs=[ValueInput(cell=cell, value='3.0')]
        )
        self.assertEqual(CellData(False, '1.0'), self.expected_values[cell])
        self.expected_values[cell] = CellData(True, '3.0')
        self.assertEqual(CellData(True, '2.0'), self.expected_values[formula_cell1])
        self.expected_values[formula_cell1] = CellData(True, '', error='Деление на 0')
        self.assertEqual(CellData(True, '4.0'), self.expected_values[formula_cell2])
        self.expected_values[formula_cell2] = CellData(True, '4.0')
        self.assertEqual(
            set(Value.objects.filter(q & Q(document=self.parent_document))),
            set(result.values)
        )
        self._test_values(self.expected_values, self.parent_document)

    def test_formula_multiple(self) -> None:
        """Тестирование изменения нескольких ячеек с зависимыми формулами."""
        form = self.forms[0]
        q = Q(column__sheet_id=form.id, column__index=2, row__index__gte=1, row__index__lte=2)
        cells = Cell.objects.filter(q)
        result = update_or_create_values(
            user=self.user,
            document=self.parent_document,
            sheet_id=form.id,
            value_inputs=[ValueInput(cell=cell, value=f'{3 + i:.1f}') for i, cell in enumerate(cells)]
        )
        for i, cell in enumerate(cells):
            self.assertEqual(CellData(False, '1.0'), self.expected_values[cell])
            self.expected_values[cell] = CellData(True, f'{3 + i:.1f}')
            formula_cell_q = Q(column__sheet_id=self.forms[0].id, column__index=4, row__index=1 + i)
            formula_cell1 = Cell.objects.get(formula_cell_q)
            q |= formula_cell_q
            self.assertEqual(CellData(True, '2.0'), self.expected_values[formula_cell1])
            self.expected_values[formula_cell1] = CellData(True, f'{4 + i:.1f}')
            formula_cell_q = Q(column__sheet_id=self.forms[1].id, column__index=4, row__index=1 + i)
            formula_cell2 = Cell.objects.get(formula_cell_q)
            q |= formula_cell_q
            self.assertEqual(CellData(True, '4.0'), self.expected_values[formula_cell2])
            self.expected_values[formula_cell2] = CellData(True, f'{8 + 2 * i:.1f}')
        self.assertEqual(
            set(Value.objects.filter(q & Q(document=self.parent_document))),
            set(result.values)
        )
        self._test_values(self.expected_values, self.parent_document)

    def test_aggregation_single(self) -> None:
        """Тестирование изменения одной ячейки с зависимой агрегацией несколько раз подряд.

        Повторное изменение проверят правильную установку `value` и `extra_value`,
        если 'Value' агрегирующей ячейки было ранее создано.
        """
        form = self.forms[0]
        document = self.children_documents[0]
        q = Q(column__sheet_id=form.id, column__index=1, row__index=3)
        cell = Cell.objects.get(q)
        for v in [3.0, 4.0, 5.0]:
            str_v = f'{v:.1f}'
            result = update_or_create_values(
                user=self.user,
                document=document,
                sheet_id=form.id,
                value_inputs=[ValueInput(cell, value=str_v)]
            )
            self.assertEqual(2, len(result.values))
            value = Value.objects.get(q & Q(document=document))
            self.assertTrue(value in result.values)
            self._test_value(value, (str_v, None))
            aggregation_value = Value.objects.get(column=self.aggregation_cell.column, row=self.aggregation_cell.row)
            self.assertTrue(aggregation_value in result.values)
            self._test_value(aggregation_value, (f'{7.0 + v:.1f}', None))

    def test_aggregation_multiple(self) -> None:
        """Тестирование изменения нескольких ячеек с зависимой агрегацией."""
        form = self.forms[0]
        document = self.children_documents[0]
        cell1_q = Q(column__sheet_id=form.id, column__index=1, row__index=3)
        cell1 = Cell.objects.get(cell1_q)
        cell2_q = Q(column__sheet_id=form.id, column__index=2, row__index=3)
        cell2 = Cell.objects.get(cell2_q)
        result = update_or_create_values(
            user=self.user,
            document=document,
            sheet_id=form.id,
            value_inputs=[ValueInput(cell1, value='4.0'), ValueInput(cell2, value='5.0')]
        )
        self.assertEqual(3, len(result.values))
        for q, v in [(cell1_q, '4.0'), (cell2_q, '5.0')]:
            value = Value.objects.get(q & Q(document=document))
            self.assertTrue(value in result.values)
            self._test_value(value, (v, None))
        aggregation_value = Value.objects.get(column=self.aggregation_cell.column, row=self.aggregation_cell.row)
        self.assertTrue(aggregation_value in result.values)
        self._test_value(aggregation_value, ('15.0', None))

    def test_aggregation_depends_on_formula(self) -> None:
        """Тестирование изменения ячейки, от которой зависит формула.

        Также существует агрегация, зависящая от этой формулы.
        """
        form = self.forms[0]
        document = self.children_documents[0]
        q = Q(column__sheet_id=form.id, column__index=2, row__index=1)
        cell = Cell.objects.get(q)
        result = update_or_create_values(
            user=self.user,
            document=document,
            sheet_id=form.id,
            value_inputs=[ValueInput(cell, value='3.0')]
        )
        self.assertEqual(4, len(result.values))
        value = Value.objects.get(q & Q(document=document))
        self.assertTrue(value in result.values)
        self._test_value(value, ('3.0', None))
        formula1_value = Value.objects.get(column__sheet_id=form.id, column__index=4, row__index=1, document=document)
        self.assertTrue(formula1_value in result.values)
        self._test_value(formula1_value, ('4.0', None))
        formula2_value = Value.objects.get(
            column__sheet_id=self.forms[1].id,
            column__index=4,
            row__index=1,
            document=document,
        )
        self.assertTrue(formula2_value in result.values)
        self._test_value(formula2_value, ('8.0', None))
        aggregation_value = Value.objects.get(
            column=self.aggregation_aggregation_depends_cell.column,
            row=self.aggregation_aggregation_depends_cell.row,
        )
        self.assertTrue(aggregation_value in result.values)
        self._test_value(aggregation_value, ('24.0', None))

    def test_formula_depends_on_aggregation(self) -> None:
        """Тестирование изменения ячейки, от которой зависит агрегация.

        Также существует формула, зависящая от этой агрегации."""
        form = self.forms[0]
        document = self.children_documents[0]
        q = Q(column__sheet_id=form.id, column__index=1, row__index=3)
        cell = Cell.objects.get(q)
        self.aggregation_formula_depends_cell.formula = '=SUM(A1:C1)'
        self.aggregation_formula_depends_cell.save(update_fields=('formula',))
        self.aggregation_cache_container.change_formula(
            f'{get_column_letter(self.aggregation_columns[3].index)}{self.aggregation_row.index}',
            self.aggregation_formula_depends_cell.formula
        )
        self.aggregation_cache_container.save()
        result = update_or_create_values(
            user=self.user,
            document=document,
            sheet_id=form.id,
            value_inputs=[ValueInput(cell, value='3.0')]
        )
        self.assertEqual(3, len(result.values))
        value = Value.objects.get(q & Q(document=document))
        self.assertTrue(value in result.values)
        self._test_value(value, ('3.0', None))
        aggregation_value = Value.objects.get(column=self.aggregation_cell.column, row=self.aggregation_cell.row)
        self.assertTrue(aggregation_value in result.values)
        self._test_value(aggregation_value, ('10.0', None))
        formula_value = Value.objects.get(
            column=self.aggregation_formula_depends_cell.column,
            row=self.aggregation_formula_depends_cell.row,
        )
        self.assertTrue(formula_value in result.values)
        self._test_value(formula_value, ('32.0', None))

    def _test_values(self, values: dict[Cell, CellData], document: Document) -> None:
        """Тестирование значений ячеек."""
        for form in self.forms:
            rows = form.rowdimension_set.order_by('index')
            columns = form.columndimension_set.order_by('index')
            for row in rows:
                for column in columns:
                    cell = Cell.objects.get(column=column, row=row)
                    cell_data = values[cell]
                    if cell_data.has_value:
                        value = Value.objects.get(column=column, row=row, document=document)
                        self._test_value(value, (cell_data.value, cell_data.error))
                    else:
                        self.assertFalse(Value.objects.filter(column=column, row=row).exists())
                        self.assertEqual(cell.default, cell_data.value)

    def _test_value(self, value: Value, data: tuple[str, str | None]) -> None:
        """Тестирование данных значения."""
        self.assertEqual(value.value, data[0])
        self.assertEqual(value.error, data[1])


@override_settings(CACHES={'default': {'BACKEND': 'django.core.cache.backends.locmem.LocMemCache'}})
class RecalculateAllCellsTestCase(TestCase):
    """Тестирование функции `recalculate_all_cells`."""

    def setUp(self) -> None:
        """Создание данных для тестирования."""
        self.superuser = User.objects.create(username='superuser', email='superuser@gmain.com', is_superuser=True)

        self.organization_content_type = ContentType.objects.get_for_model(Organization)
        self.project = Project.objects.create(content_type=self.organization_content_type)
        self.period = Period.objects.create(project=self.project)

        self.forms: list[Sheet] = []
        self.columns: list[ColumnDimension] = []
        for i in range(1, 3):
            form = Sheet.objects.create(name=f'Форма №{i}', period=self.period, show_head=False)
            row = RowDimension.objects.create(index=1, sheet=form)
            columns = [ColumnDimension.objects.create(index=i, sheet=form) for i in range(1, 4)]
            self.columns.extend(columns)
            form_cache_container = SheetFormulaContainerCache(name=form.name)
            for j, column in enumerate(columns, 1):
                if j == 1:
                    formula = "=SUM(B1:C1)"
                    Cell.objects.create(
                        kind=KindCell.NUMERIC,
                        formula=formula,
                        default=f'5.0',
                        column=column,
                        row=row,
                    )
                    form_cache_container.add_formula('A1', formula)
                else:
                    Cell.objects.create(
                        kind=KindCell.NUMERIC,
                        default=f'{j:.1f}',
                        column=column,
                        row=row,
                    )
            form_cache_container.save(sheet_id=form.id)
            self.forms.append(form)

        form3 = Sheet.objects.create(name=f'Форма №3', period=self.period, show_head=False)
        row = RowDimension.objects.create(index=1, sheet=form3)
        column = ColumnDimension.objects.create(index=1, sheet=form3)
        self.columns.append(column)
        form3_cache_container = SheetFormulaContainerCache(name=form3.name)
        form3_formula = "='Форма №1'!A1 + 'Форма №2'!A1"
        Cell.objects.create(
            kind=KindCell.NUMERIC,
            formula=form3_formula,
            default=f'10.0',
            column=column,
            row=row,
        )
        form_3_aggregate_column = ColumnDimension.objects.create(index=2, sheet=form3)
        form3_aggregate_cell =  Cell.objects.create(
            kind=KindCell.NUMERIC,
            default=f'15.0',
            column=form_3_aggregate_column,
            row=row,
        )
        form3_cache_container.add_formula('A1', form3_formula)
        form3_cache_container.save(sheet_id=form3.id)
        self.forms.append(form3)

        aggregation_form = Sheet.objects.create(name='Форма агрегации', period=self.period, show_child=False)
        row = RowDimension.objects.create(index=1, sheet=aggregation_form)
        self.aggregation_column = ColumnDimension.objects.create(index=1, sheet=aggregation_form)
        aggregation_cell = Cell.objects.create(
            kind=KindCell.NUMERIC,
            aggregation=True,
            default=f'0.0',
            column=self.aggregation_column,
            row=row,
        )
        aggregation_cell.to_cells.create(from_cell=form3_aggregate_cell)
        self.forms.append(aggregation_form)

        parent_organization = Organization.objects.create(attributes='')
        self.period.division_set.create(object_id=parent_organization.id)
        self.parent_document = Document.objects.create(period=self.period, object_id=parent_organization.id)
        self.parent_document.sheets.set(self.forms)

        self.documents: list[Document] = []
        for i in range(2):
            organization = Organization.objects.create(attributes='', parent=parent_organization)
            self.period.division_set.create(object_id=organization.id)
            document = Document.objects.create(period=self.period, object_id=organization.id)
            document.sheets.set(self.forms)
            form = self.forms[0] if i == 0 else self.forms[1]
            Value.objects.create(
                document=document,
                sheet=form,
                column=form.columndimension_set.last(),
                row=form.rowdimension_set.last(),
                value=f'{10 + i:.1f}',
            )
            self.documents.append(document)

    def test_recalculate_all_cells(self) -> None:
        """Тестирование функции `recalculate_all_cells`."""
        self._test_values(self.documents[0], self.columns, (
            ('5.0', False),
            ('2.0', False),
            ('10.0', True),
            ('5.0', False),
            ('2.0', False),
            ('3.0', False),
            ('10.0', False),
        ))
        self._test_values(self.documents[1], self.columns, (
            ('5.0', False),
            ('2.0', False),
            ('3.0', False),
            ('5.0', False),
            ('2.0', False),
            ('11.0', True),
            ('10.0', False),
        ))
        self._test_values(self.parent_document, [self.aggregation_column], (('0.0', False),))
        recalculate_all_cells(self.superuser, self.period)
        self._test_values(self.documents[0], self.columns, (
            ('12.0', True),
            ('2.0', False),
            ('10.0', True),
            ('5.0', True),
            ('2.0', False),
            ('3.0', False),
            ('17.0', True),
        ))
        self._test_values(self.documents[1], self.columns, (
            ('5.0', True),
            ('2.0', False),
            ('3.0', False),
            ('13.0', True),
            ('2.0', False),
            ('11.0', True),
            ('18.0', True),
        ))
        self._test_values(self.parent_document, [self.aggregation_column], (('30.0', True),))

    def _test_values(
        self,
        document: Document,
        columns: list[ColumnDimension],
        values: Iterable[tuple[str, bool]]
    ) -> None:
        """Тестирование значений ячеек."""
        for column, value_exist in zip(columns, values):
            expected_value, exist = value_exist
            value = Value.objects.filter(document=document, column=column).first()
            self.assertEqual(exist, bool(value))
            actual_value = value.value if value else Cell.objects.filter(column=column).first().default
            self.assertEqual(expected_value, actual_value)
