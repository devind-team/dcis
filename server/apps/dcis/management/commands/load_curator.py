"""Модуль с командой."""
import datetime
from os.path import join

from devind_core.models import Profile, ProfileValue
from devind_dictionaries.models import Organization
from devind_helpers.orm_utils import get_object_or_none
from django.core.management.base import BaseCommand
from django.utils.timezone import make_aware
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from apps.core.models import User
from apps.dcis.models import CuratorGroup
from devind.settings import STATICFILES_DIRS


class Command(BaseCommand):
    """Команда."""

    helps = 'Первоначальное развертывание проекта.'

    @classmethod
    def change_organization(cls, ) -> Organization:
        name_organization = 'МИНИСТЕРСТВО НАУКИ И ВЫСШЕГО ОБРАЗОВАНИЯ РОССИЙСКОЙ ФЕДЕРАЦИИ'
        organization_id = Organization.objects.get(name=name_organization).id

        if organization_id != 0:
            Organization.objects.filter(pk=organization_id).update(id=0)

        return Organization.objects.get(pk=0)

    @classmethod
    def check_user(
        cls,
        row: tuple
    ) -> int:
        organization = cls.change_organization()
        user = get_object_or_none(
            User,
            profilevalue__profile__code='uuid',
            profilevalue__value=row[2].value
        )
        if user is None:
            user = User.objects.create(
                username=row[3].value,
                last_name=row[4].value,
                first_name=row[5].value,
                sir_name=row[6].value,
                email=row[7].value
            )
            ProfileValue.objects.create(
                profile=Profile.objects.get(code='uuid'),
                user=user,
                value=row[2].value
            )
        user.agreement = make_aware(datetime.datetime.now())
        user.save(update_fields=('agreement',))
        organization.users.add(user.id)
        return user.id

    def handle(self, *args, **options) -> None:
        """Описание команды."""

        workbook: Workbook = load_workbook(filename=join(STATICFILES_DIRS[1], 'kurators_orgs_users.xlsx'))
        sheets: list[str] = workbook.sheetnames

        kur_groups: Worksheet = workbook[sheets[0]]
        users_id: list[int] = []
        curator_groups: list[str] = []
        for i, row in enumerate(kur_groups):
            if i == 0:
                continue
            users_id.append(self.check_user(row=row))
            curator_groups.append(row[1].value)

        for group_name in zip(curator_groups, users_id):
            group: CuratorGroup = CuratorGroup.objects.create(name=group_name[0].replace('  ', ' '))
            group.users.add(group_name[1])

        kur_groups_orgs: Worksheet = workbook[sheets[1]]
        for i, row in enumerate(kur_groups_orgs):
            if i == 0:
                continue
            group: CuratorGroup = CuratorGroup.objects.get(name=row[0].value.replace('  ', ' '))
            group.organization.add(row[1].value)
